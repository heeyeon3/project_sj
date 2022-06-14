from copyreg import constructor
from datetime import datetime, timedelta
from email.policy import default
from flask_jwt_extended import jwt_required
from flask_login import current_user
from flask_restful import Resource, reqparse, request
import openpyxl
from pymysql import NULL
from sqlalchemy import false, true
from models.alarm import AlarmModel
from models.datarogger import DataroggerModel
from models.sensordetail import SensorDetailModel
from resource.datarogger import DataloggerList, Datarogger
from resource.sensordetail import Sensordetailmodal
from utils.jsonutil import AlchemyEncoder as jsonEncoder
import json
from models.editdata import EditdataModel
from models.sensor import SensorModel
from models.sensorgroup import SensorgroupModel
from models.fuction import FunctionModel
from flask import send_from_directory
import requests # excel backup
from openpyxl    import load_workbook, Workbook
from openpyxl.styles    import PatternFill
from config.properties import *
import pandas as pd
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from resource.log import LogMessage

from math import pi, sin, cos, tan, log, e
import math
import pandas as pd

import logging

from datetime import datetime, timedelta
import copy

class Editdata(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_name', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_data_date', type=str)
    parse.add_argument('sensor_data', type=str)


    parse.add_argument('sensor_data_x', type=str)
    parse.add_argument('sensor_data_y', type=str)

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('sensor_id_x', type=str)
    parse.add_argument('sensor_id_y', type=str)
    # parse.add_argument('date_time_end', type=str)
    # parse.add_argument('intervalday', type=str)
    # parse.add_argument('time', type=str)
    # parse.add_argument('sensor_name', type=str)


    def post(self):

        params = Editdata.parse.parse_args()
        # print(params)

  

        sensor_id = params["sensor_id"]

        sensor_obj = SensorModel.find_by_id(sensor_id)
        sensor_name = sensor_obj.sensor_name
        datarogger_id = str(sensor_obj.datarogger_id)
        # sensor_name = params["sensor_name"]
        # datarogger_id = params["datarogger_id"]
        sensor_data_date = params["sensor_data_date"]
        sensor_data = params["sensor_data"]

        datarogger_obj = DataroggerModel.find_by_id(datarogger_id)
        project_id = datarogger_obj.project_id

        editdata_obj = EditdataModel.find_by_editdata(sensor_name, datarogger_id, sensor_data_date)

        edit_sensordata = editdata_obj.sensor_data
        # print(edit_sensordata)

        sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_id)
        sensor_initial_date = sensordetail_obj.sensor_initial_date
        sensor_initial_date_2 = str(sensor_initial_date)[0:4]+"."+str(sensor_initial_date)[5:7]+"."+str(sensor_initial_date)[8:10] +" "+str(sensor_initial_date)[11:19]
  

       

        user_id = current_user.user_id
        modify_date = datetime.now() 
        try:
            editdata_obj = EditdataModel.find_by_editdata(sensor_name, datarogger_id, sensor_data_date)

            edit_sensordata = editdata_obj.sensor_data
            editdata_obj.sensor_data = sensor_data
            editdata_obj.edit_yn = 'Y'
            editdata_obj.user_id = user_id
            editdata_obj.modify_date = modify_date
            editdata_obj.save_to_db()

            message = "project_id: "+str(project_id)+ " sensor_id: "+sensor_id+ " sensor_name: " +sensor_name +" datarogger_id: " +datarogger_id +" "+sensor_data_date+" 시간 데이터를 "+sensor_data +"로 수정하였습니다."

            LogMessage.sensor_edit("data_edit", message, "0502")


            # 수정 값이 초기값과 같을때 센서에 등록된 초기값 변경
            if sensor_initial_date_2 == sensor_data_date:
                # print("same")
                sensordetail_obj.sensor_initial_data = sensor_data
                sensordetail_obj.save_to_db()

            alarm_obj = AlarmModel.find_by_sensor_id_datetime(datarogger_id, sensor_id, str(sensor_data_date))
            if alarm_obj is not None:
             
                alarm_obj.use_yn = 'N'
                alarm_obj.alarm_status = 'Y'
                alarm_obj.save_to_db()
            
       

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500


        return {'resultCode': '0', "resultString": " 등록 되었습니다."}, 200

    #스캐터일때 데이터 수정
    def put(self):

        params = Editdata.parse.parse_args()

  

        sensor_idx = params["sensor_id_x"]
        sensor_idy = params["sensor_id_y"]

        sensor_obj_x = SensorModel.find_by_id(sensor_idx)
        sensor_name_x = sensor_obj_x.sensor_name
        datarogger_id_x = str(sensor_obj_x.datarogger_id)

        sensor_obj_y = SensorModel.find_by_id(sensor_idy)
        sensor_name_y = sensor_obj_y.sensor_name
        datarogger_id_y = str(sensor_obj_y.datarogger_id)
        
        sensor_data_date = params["sensor_data_date"]
        sensor_data_x = params["sensor_data_x"]
        sensor_data_y = params["sensor_data_y"]

        editdata_obj_x = EditdataModel.find_by_editdata(sensor_name_x, datarogger_id_x, sensor_data_date)
        editdata_obj_y = EditdataModel.find_by_editdata(sensor_name_y, datarogger_id_y, sensor_data_date)

        edit_sensordata_x = editdata_obj_x.sensor_data
        edit_sensordata_y = editdata_obj_y.sensor_data
     
        user_id = current_user.user_id
        modify_date = datetime.now() 

        sensordetail_obj_x = SensorDetailModel.find_by_sensor_id(sensor_idx)
        sensordetail_obj_y = SensorDetailModel.find_by_sensor_id(sensor_idy)
        sensor_initial_date = sensordetail_obj_x.sensor_initial_date
        sensor_initial_date_2 = str(sensor_initial_date)[0:4]+"."+str(sensor_initial_date)[5:7]+"."+str(sensor_initial_date)[8:10] +" "+str(sensor_initial_date)[11:19]

        datarogger_obj = DataroggerModel.find_by_id(datarogger_id_x)
        project_id = datarogger_obj.project_id
        try:
            editdata_obj_x = EditdataModel.find_by_editdata(sensor_name_x, datarogger_id_x, sensor_data_date)
            editdata_obj_y = EditdataModel.find_by_editdata(sensor_name_y, datarogger_id_y, sensor_data_date)

            
            editdata_obj_x.sensor_data = sensor_data_x
            editdata_obj_x.edit_yn = 'Y'
            editdata_obj_x.user_id = user_id
            editdata_obj_x.modify_date = modify_date

            message = "project_id: "+str(project_id)+ " sensor_id: "+sensor_idx+ " sensor_name: " +sensor_name_x +" datarogger_id: " +datarogger_id_x +" "+sensor_data_date+" 시간 데이터를 "+sensor_data_x +"로 수정하였습니다."
            LogMessage.sensor_edit("data_edit", message, "0502")

            editdata_obj_y.sensor_data = sensor_data_y
            editdata_obj_y.edit_yn = 'Y'
            editdata_obj_y.user_id = user_id
            editdata_obj_y.modify_date = modify_date

            message = "project_id: "+str(project_id)+ " sensor_id: "+sensor_idy+ " sensor_name: " +sensor_name_y +" datarogger_id: " +datarogger_id_y +" "+sensor_data_date+" 시간 데이터를 "+sensor_data_y +"로 수정하였습니다."
            LogMessage.sensor_edit("data_edit", message, "0502")

            editdata_obj_x.save_to_db()
            editdata_obj_y.save_to_db()

            if sensor_initial_date_2 == sensor_data_date:
                sensordetail_obj_x.sensor_initial_data = sensor_data_x
                sensordetail_obj_y.sensor_initial_data = sensor_data_y

                sensordetail_obj_x.save_to_db()
                sensordetail_obj_y.save_to_db()


        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500


        return {'resultCode': '0', "resultString": " 등록 되었습니다."}, 200


class editdatatable(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensor_name', type=str)

    def get(self):
   
        
        date_time_start =              request.args.get('date_time_start')
        date_time_end =              request.args.get('date_time_end')
        intervalday =              request.args.get('intervalday')
        time =              request.args.get('time')
        sensor_id =              request.args.get('sensor_id')

        print(intervalday, time)

        sensor_obj = SensorModel.find_by_id(sensor_id)
        datarogger_id = str(sensor_obj.datarogger_id)
        sensor_name = sensor_obj.sensor_name

        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_id)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']
        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'

        param = (sensor_name, datarogger_id, date_time_start, date_time_end, '', time )
        data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid_useyn_all(param)]


        if sensor_fx1 and len(sensor_fx1) != 0 :
            for i in range(len(data)):
               
                current_fx = sensor_fx1
              
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                result1 = eval(current_fx)
            

                data[i]['sensor_fx1_data']=str(round(result1,4))
              
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            for i in range(len(data)):
                
                current_fx = sensor_fx2
               
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                result1 = eval(current_fx)
               

                data[i]['sensor_fx2_data']=str(round(result1,4))
             

        if sensor_fx3 and len(sensor_fx3) != 0 :
            for i in range(len(data)):
                
                current_fx = sensor_fx3
                
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
               

                data[i]['sensor_fx3_data']=str(round(result1,4))
            

        if sensor_fx4 and len(sensor_fx4) != 0 :
            for i in range(len(data)):
               
                current_fx = sensor_fx4
                
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
               

                data[i]['sensor_fx4_data']=str(round(result1,4))
          

        if sensor_fx5 and len(sensor_fx5) != 0 :
            for i in range(len(data)):
                
                current_fx = sensor_fx5
             
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
               

                data[i]['sensor_fx5_data']=str(round(result1,4))
              



        return {'resultCode': '0', "resultString": "SUCCESS" ,"data":data}, 200




    def post(self):

  
        params = editdatatable.parse.parse_args()
        print("================================================================================================================================================================")
       
      
        sensor_id = params["sensor_id"]
        sensor_name = params["sensor_name"]
        # datarogger_id = params["datarogger_id"]
        date_time_start = params["date_time_start"]
        date_time_end = params["date_time_end"]
        intervalday = params["intervalday"]
        time = params["time"]

        print(params)
 

        sensor_obj = SensorModel.find_by_id(sensor_id)
        datarogger_id = str(sensor_obj.datarogger_id)
        sensor_name = sensor_obj.sensor_name

        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_id)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        # sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        # sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        # sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        # sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        # sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        # sensor_fx_check = sensorlist[0]['sensor_fx_check']
        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']
        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'

      
        format = "%Y-%m-%d"
        start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
        end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")

        print(start, end)
        realdata = []
        currentIndex = 0

       
        dates = []
        dates = pd.date_range(start=start, end=end, freq='1d')
        dates = dates.sort_values(ascending = False)
        sortdateall = []

        for i in range(len(dates)):
            date_i = str(dates[i])
            sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])
            
    
        param = (sensor_name, datarogger_id, date_time_start, date_time_end, '', time )
        if  len(time) ==0 and len(intervalday) !=0:
            freq = intervalday+'d'
            sortdate = []
            realdata = pd.date_range(start=start, end=end, freq=freq)
            realdata = realdata.sort_values(ascending = False)
            # print(realdata)
            # for i in range(len(realdata)):
            #     date_i = str(realdata[i])
            #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

            sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
            befor_days = datetime.now()
            day_i = 1
            while start < befor_days:
                befor_days = end - timedelta(days= day_i * int(intervalday))
                befor_one_days = str(befor_days)
                sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                day_i += 1
                print(befor_days)
         
            sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
            
            data = []   
            for i in range(len(sensordatas)):
                currnetdate = str(sensordatas[i]['sensor_data_date'])
              
                totaldata = 0
                totallen = 0
                if currnetdate in sortdate:
                    # currentdaatime = currnetdate[11:20]
                    dateindex = sortdateall.index(currnetdate[0:10])
                  
                    for idx in range(int(intervalday)):
                        currentIndex = dateindex+idx

                        if currentIndex < len(sortdateall):
                            lastdate = sortdateall[currentIndex] 
                          
                            for j in range(len(sensordatas)):
                                if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                    totaldata += float(sensordatas[j]['sensor_data'])
                                    totallen += 1
                                
                                    break
  

                        if idx == int(intervalday)-1:
                            avg_data = round(totaldata/totallen, 4)
                            data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name, 'sensor_data': str(avg_data)})


        elif intervalday !='1' and len(intervalday) !=0:
            freq = intervalday+'d'
            sortdate = []
            realdata = pd.date_range(start=start, end=end, freq=freq)
            realdata = realdata.sort_values(ascending = False)

            sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
            befor_days = datetime.now()
            day_i = 1
            while start < befor_days:
                befor_days = end - timedelta(days= day_i * int(intervalday))
                befor_one_days = str(befor_days)
                sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                day_i += 1



            sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
   

            data = []
            for i in range(len(sensordatas)):
                currnetdate = str(sensordatas[i]['sensor_data_date'])
               
                totaldata = 0
                totallen = 0
                if currnetdate[0:10] in sortdate:
                    currentdaatime = currnetdate[11:20]
                    dateindex = sortdateall.index(currnetdate[0:10])
                  
                    for idx in range(int(intervalday)):
                        currentIndex = dateindex+idx

                        if currentIndex < len(sortdateall):
                            lastdate = sortdateall[currentIndex] +" "+ currentdaatime
                          
                            for j in range(len(sensordatas)):
                                if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                    totaldata += float(sensordatas[j]['sensor_data'])
                                    totallen += 1
                                
                                    break
  

                        if idx == int(intervalday)-1:
                            avg_data = round(totaldata/totallen, 4)
                            data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})


        else:
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
           

 
        if sensor_fx1 and len(sensor_fx1) != 0 :
            for i in range(len(data)):
               
                current_fx = sensor_fx1
              
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                result1 = eval(current_fx)
            

                data[i]['sensor_fx1_data']=str(round(result1,4))
              
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            for i in range(len(data)):
                
                current_fx = sensor_fx2
               
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                result1 = eval(current_fx)
               

                data[i]['sensor_fx2_data']=str(round(result1,4))
             

        if sensor_fx3 and len(sensor_fx3) != 0 :
            for i in range(len(data)):
                
                current_fx = sensor_fx3
                
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
               

                data[i]['sensor_fx3_data']=str(round(result1,4))
            

        if sensor_fx4 and len(sensor_fx4) != 0 :
            for i in range(len(data)):
               
                current_fx = sensor_fx4
                
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
               

                data[i]['sensor_fx4_data']=str(round(result1,4))
          

        if sensor_fx5 and len(sensor_fx5) != 0 :
            for i in range(len(data)):
                
                current_fx = sensor_fx5
             
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
               

                data[i]['sensor_fx5_data']=str(round(result1,4))
              



        return {'resultCode': '0', "resultString": "SUCCESS" ,"data":data}, 200



class editdatatableall(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('sensorgroup_type', type=str)

    def post(self):

        print("Sensorgroupedittable")
        params = editdatatableall.parse.parse_args()
      

        sensorgroup_id = params["sensorgroup_id"]
        datarogger_id = params["datarogger_id"]
        date_time_start = params["date_time_start"]
        date_time_end = params["date_time_end"]
        intervalday = params["intervalday"]
        time = params["time"]
        sensorgroup_type = params["sensorgroup_type"]
      
        sensordata_list = [dict(r) for r in SensorModel.find_by_datarogger_id_snesorgroup(sensorgroup_id)]
    
        totaldatalist = []
        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_initial_date = sensorgroup_obj.sensorgroup_initial_date

        
        sensor_gauge_factor = sensorgroup_obj.sensorgroup_initial_date

        #sensor date list
        datelist = []
        for i in range(len(sensordata_list)):
            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']

            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            if len(datelist) < len(data):
                datelist = []
                for i in range(len(data)):
                    datelist.append({"sensor_data_date":data[i]['sensor_data_date']})
        
            

        for i in range(len(sensordata_list)):
            # print(sensordata_list[i]['sensor_name'])

            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']

         
            if sensor_initial_data is None or len(sensor_initial_data) == 0:
                sensor_initial_data = '0'
            if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
                sensor_gauge_factor = '0'

      

            sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(str(sensor_id))]

            # print("sensorlist", sensorlist)

            if sensorlist:
                sensor_fx1 = sensorlist[0]['sensor_fx1']
                sensor_fx2 = sensorlist[0]['sensor_fx2']
                sensor_fx3 = sensorlist[0]['sensor_fx3']
                sensor_fx4 = sensorlist[0]['sensor_fx4']
                sensor_fx5 = sensorlist[0]['sensor_fx5']
                sensor_fx_check = sensorlist[0]['sensor_fx_check']


                # sensor_initial_data = sensorlist[0]['sensor_initial_data']

                sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
                sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
                sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
                sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
                sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
          

                param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

     
                format = "%Y-%m-%d"
                start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
                end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")
        

                dates = []
                dates = pd.date_range(start=start, end=end, freq='1d')
                dates = dates.sort_values(ascending = False)
                sortdateall = []

                for i in range(len(dates)):
                    date_i = str(dates[i])
                    sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])


                if  len(time) ==0 and len(intervalday) !=0:
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)
                    # print(realdata)
                    # for i in range(len(realdata)):
                    #     date_i = str(realdata[i])
                    #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1
                
                    sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
                    
                    data = []   
                    for i in range(len(sensordatas)):
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate in sortdate:
                            # currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx

                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] 
                                
                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                        
                                            break

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name,  'datarogger_id': datarogger_id, 'sensor_data': str(avg_data)})


  
                elif len(intervalday) !=0 and int(intervalday) >1:
                    realdata = []
                    currentIndex=0
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1

                    sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
   

                    data = []

                    for i in range(len(sensordatas)):
              
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate[0:10] in sortdate:
                           
                            currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                            

                            for idx in range(int(intervalday)):
                            
                                # if idx != 0:
                                
                                currentIndex = dateindex+idx
                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] +" "+ currentdaatime
                                    # print("lastdate", lastdate)

                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                            break
                        
                                            


                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                 
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'],  'datarogger_id': datarogger_id, 'sensor_data': str(avg_data)})

                 

                else:
               
                   
                    # data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                    data01 = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                    datalist01 = copy.deepcopy(datelist)
                
                    for i in range(len(datalist01)):
                        datayn = 0
                        for j in range(len(data01)):
                            
                            if datalist01[i]['sensor_data_date'] == data01[j]['sensor_data_date']:
                                datalist01[i]['sensor_name'] = data01[j]['sensor_name']
                                datalist01[i]['datarogger_id'] = data01[j]['datarogger_id']
                                datalist01[i]['sensor_data'] = data01[j]['sensor_data']
                                datalist01[i]['sensor_fx_check_data'] = None
                                datayn =1
                                break

                        if datayn == 0:
                            datalist01[i]['sensor_name'] = sensor_name
                            datalist01[i]['datarogger_id'] = datarogger_id
                            datalist01[i]['sensor_data'] = None
                            datalist01[i]['sensor_fx_check_data'] = None


                    data = copy.deepcopy(datalist01)

     

                if sensor_fx1 and len(sensor_fx1) != 0 :
                  
                    for i in range(len(data)):
                      
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx1
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)
                        

                            data[i]['sensor_fx1_data']=str(round(result1,4))

                            if sensor_fx_check == '1':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))

                        else:
                           
                            data[i]['sensor_fx1_data']=None
                 
            
                
                if sensor_fx2 and len(sensor_fx2) != 0 :
                   
                    for i in range(len(data)):

                        if data[i]['sensor_data']:
                            current_fx = sensor_fx2
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx2_data']=str(round(result1,4))

                            if sensor_fx_check == '2':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))

                        else:
                            data[i]['sensor_fx2_data']=None
           
                if sensor_fx3 and len(sensor_fx3) != 0 :
                  
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx3
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx3_data']=str(round(result1,4))

                            if sensor_fx_check == '3':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))
                        else:
                            data[i]['sensor_fx3_data']=None
                        
              

                if sensor_fx4 and len(sensor_fx4) != 0 :
                    
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx4
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx4_data']=str(round(result1,4))

                            if sensor_fx_check == '4':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))
                        else:

                            data[i]['sensor_fx4_data']=None
              

                if sensor_fx5 and len(sensor_fx5) != 0 :
                
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                    
                            current_fx = sensor_fx5
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)
                        

                            data[i]['sensor_fx5_data']=str(round(result1,4))

                            if sensor_fx_check == '5':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))

                        else:
                            data[i]['sensor_fx5_data']=None
              

                totaldatalist.append(data)
        
        # newdaylist = []
        # for i in range(len(totaldatalist)):
        #     if len(newdaylist) < len(totaldatalist[i]):
        #         for j in range(len(totaldatalist[i])):
        #             newdaylist.append(totaldatalist[i][j]['sensor_data_date'])
        # newdata = []
        # for i in range(len(totaldatalist)):


             
        return {'resultCode': '0', "resultString": "SUCCESS", "data": totaldatalist}, 200




class editdatatableallgraph(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('sensorgroup_type', type=str)

    def post(self):

        print("Sensorgroupedittable")
        params = editdatatableall.parse.parse_args()
      
        sensorgroup_id = params["sensorgroup_id"]
        datarogger_id = params["datarogger_id"]
        date_time_start = params["date_time_start"]
        date_time_end = params["date_time_end"]
        intervalday = params["intervalday"]
        time = params["time"]
        sensorgroup_type = params["sensorgroup_type"]
        # print("sensorgroup_type", sensorgroup_type)

        sensordata_list = [dict(r) for r in SensorModel.find_by_datarogger_id_snesorgroup(sensorgroup_id)]
   

        totaldatalist = []
        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_initial_date = sensorgroup_obj.sensorgroup_initial_date
        sensorgroup_interval = sensorgroup_obj.sensorgroup_interval

        datelist = []

        for i in range(len(sensordata_list)):
        
            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_fx_check = sensordata_list[i]['sensor_fx_check']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']
        
            if sensor_initial_data is None or len(sensor_initial_data) == 0:
                sensor_initial_data = '0'
            if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
                sensor_gauge_factor = "0"

            sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(str(sensor_id))]

            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            if len(datelist) < len(data):
                for j in range(len(data)):
                    datelist.append({"sensor_data_date":data[j]['sensor_data_date']})


        for i in range(len(sensordata_list)):
        
            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_fx_check = sensordata_list[i]['sensor_fx_check']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']
        
            if sensor_initial_data is None or len(sensor_initial_data) == 0:
                sensor_initial_data = '0'
            if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
                sensor_gauge_factor = "0"

            # if sensorgroup_initial_date:
            #     sensor_initial_data_obj = [dict(r) for r in EditdataModel.sensor_initail_data_sensorgroup(str(sensor_id), str(sensorgroup_initial_date))]
            #     print("sensor_initial_data_obj", sensor_initial_data_obj)

            #     if len(sensor_initial_data_obj) !=0:
            #         sensor_initial_data = sensor_initial_data_obj[0]['sensor_data']

            #     else:
            #         sensor_initial_data = '0'

            # else:
            #     sensor_initial_data = '0'

            sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(str(sensor_id))]

            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            # if len(datelist) < len(data):
            #     for j in range(len(data)):
            #         datelist.append({"sensor_data_date":data[j]['sensor_data_date']})
            
            

      
            if sensorlist:
                sensor_fx1 = sensorlist[0]['sensor_fx1']
                sensor_fx2 = sensorlist[0]['sensor_fx2']
                sensor_fx3 = sensorlist[0]['sensor_fx3']
                sensor_fx4 = sensorlist[0]['sensor_fx4']
                sensor_fx5 = sensorlist[0]['sensor_fx5']


                # sensor_initial_data = sensorlist[0]['sensor_initial_data']

                sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
                sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
                sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
                sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
                sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
                # if sensor_initial_data is None:
                #     sensor_initial_data = '0'


                param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, intervalday, time )

               
                data = []
                data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

                format = "%Y-%m-%d"
                start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
                end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")
        

                dates = []
                dates = pd.date_range(start=start, end=end, freq='1d')
                dates = dates.sort_values(ascending = False)
                sortdateall = []

                for j in range(len(dates)):
                    date_i = str(dates[j])
                    sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

                if  len(time) ==0 and len(intervalday) !=0:
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)
                    # print(realdata)
                    # for i in range(len(realdata)):
                    #     date_i = str(realdata[i])
                    #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1
                        print(befor_days)
                
                    sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
                    
                    data = []   
                    for i in range(len(sensordatas)):
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate in sortdate:
                            # currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx

                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] 
                                
                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                        
                                            break
        

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name,'datarogger_id': datarogger_id, 'sensor_data': str(avg_data)})

            
              
                
                elif len(intervalday) !=0 and int(intervalday) > 1:
                    realdata = []
                    currentIndex=0
                    


                    sortdate = []


                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1
                        print(befor_days)
                
                    sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                    
                    data = []   
                    for i in range(len(sensordatas)):
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate in sortdate:
                            # currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx

                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] 
                                
                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                        
                                            break
        

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name,'datarogger_id': datarogger_id, 'sensor_data': str(avg_data)})

            



                    

                else:
                    # data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
               
                    data01 = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                    datalist01 = copy.deepcopy(datelist)
                    for i in range(len(datalist01)):
                        datayn = 0
                        for j in range(len(data01)):
                            
                            if datalist01[i]['sensor_data_date'] == data01[j]['sensor_data_date']:
                                datalist01[i]['sensor_name'] = data01[j]['sensor_name']
                                datalist01[i]['sensor_data'] = data01[j]['sensor_data']
                                datayn =1
                                break

                        if datayn == 0:
                            datalist01[i]['sensor_name'] = sensor_name
                            datalist01[i]['sensor_data'] = None

                        # print("datelist",datelist)

                    data = copy.deepcopy(datalist01)

                if sensor_fx_check == '1':
                    

                    if sensor_fx1 and len(sensor_fx1) != 0 :
                    
                        for i in range(len(data)):
                            if data[i]['sensor_data']:
                                current_fx = sensor_fx1
                                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                                current_fx = current_fx.replace('$ini', sensor_initial_data)
                                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                                result1 = eval(current_fx)
                            

                                data[i]['sensor_fx_data']=str(round(result1,4))

                            else:
                                data[i]['sensor_fx_data']=None
             
                elif sensor_fx_check == '2':
                    if sensor_fx2 and len(sensor_fx2) != 0 :
                    
                        for i in range(len(data)):

                            if data[i]['sensor_data']:
                                current_fx = sensor_fx2
                                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                                current_fx = current_fx.replace('$ini', sensor_initial_data)
                                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                                result1 = eval(current_fx)

                                data[i]['sensor_fx_data']=str(round(result1,4))

                            else:
                                data[i]['sensor_fx_data']=None


                elif sensor_fx_check == '3':
               
                    if sensor_fx3 and len(sensor_fx3) != 0 :
                
                        for i in range(len(data)):
                            if data[i]['sensor_data']:
                                current_fx = sensor_fx3
                                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                                current_fx = current_fx.replace('$ini', sensor_initial_data)
                                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                                result1 = eval(current_fx)

                                data[i]['sensor_fx_data']=str(round(result1,4))
                            else:
                                data[i]['sensor_fx_data']=None
                
                elif sensor_fx_check == '4':

                    if sensor_fx4 and len(sensor_fx4) != 0 :
                    
                        for i in range(len(data)):
                            if data[i]['sensor_data']:
                                current_fx = sensor_fx4
                                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                                current_fx = current_fx.replace('$ini', sensor_initial_data)
                                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                                result1 = eval(current_fx)

                                data[i]['sensor_fx_data']=str(round(result1,4))
                            else:

                                data[i]['sensor_fx_data']=None

                elif sensor_fx_check == '5':    

                    if sensor_fx5 and len(sensor_fx5) != 0 :
                    
                        for i in range(len(data)):
                        
                            if data[i]['sensor_data']:
                    
                                current_fx = sensor_fx5
                                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                                current_fx = current_fx.replace('$ini', sensor_initial_data)
                                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                                result1 = eval(current_fx)
                            

                                data[i]['sensor_fx_data']=str(round(result1,4))

                            else:
                                data[i]['sensor_fx_data']=None
              
                print("data", data)        
             
                totaldatalist.append(data)
        
  

        return {'resultCode': '0', "resultString": "SUCCESS", "data": totaldatalist}, 200


class exceldown(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensor_name', type=str)

    def get(self):

        print("들어옴!!!!!!!!!!!!!!!!")

        # params = editdatatable.parse.parse_args()

        # sensor_name = params["sensor_name"]
        # datarogger_id = params["datarogger_id"]
        # date_time_start = params["date_time_start"]
        # date_time_end = params["date_time_end"]
        # intervalday = params["intervalday"]
        # time = params["time"]
 

        sensor_name =              request.args.get('sensor_name')
        
        date_time_start =              request.args.get('date_time_start')
        date_time_end =              request.args.get('date_time_end')
        intervalday =              request.args.get('intervalday')
        time =              request.args.get('time')
        sensor_id =              request.args.get('sensor_id')


        sensor_obj = SensorModel.find_by_id(sensor_id)
        sensor_display_name = sensor_obj.sensor_display_name
        datarogger_id = sensor_obj.datarogger_id
        sensor_name = sensor_obj.sensor_name
       
     
        # param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, intervalday, time )

        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_id)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']
        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'

       
        now = datetime.now()
        book = Workbook()
     
        file_name = sensor_display_name +'-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

        format = "%Y-%m-%d"
        start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
        end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")

        realdata = []
        currentIndex = 0

       
        dates = []
        dates = pd.date_range(start=start, end=end, freq='1d')
        dates = dates.sort_values(ascending = False)
        sortdateall = []

        for i in range(len(dates)):
            date_i = str(dates[i])
            sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])
            
    
        param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
        
        if  len(time) ==0 and len(intervalday) !=0:
            freq = intervalday+'d'
            sortdate = []
            realdata = pd.date_range(start=start, end=end, freq=freq)
            realdata = realdata.sort_values(ascending = False)
            # print(realdata)
            # for i in range(len(realdata)):
            #     date_i = str(realdata[i])
            #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

            sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
            befor_days = datetime.now()
            day_i = 1
            while start < befor_days:
                befor_days = end - timedelta(days= day_i * int(intervalday))
                befor_one_days = str(befor_days)
                sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                day_i += 1
                print(befor_days)
         
            sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
            
            data = []   
            for i in range(len(sensordatas)):
                currnetdate = str(sensordatas[i]['sensor_data_date'])
              
                totaldata = 0
                totallen = 0
                if currnetdate in sortdate:
                    # currentdaatime = currnetdate[11:20]
                    dateindex = sortdateall.index(currnetdate[0:10])
                  
                    for idx in range(int(intervalday)):
                        currentIndex = dateindex+idx

                        if currentIndex < len(sortdateall):
                            lastdate = sortdateall[currentIndex] 
                          
                            for j in range(len(sensordatas)):
                                if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                    totaldata += float(sensordatas[j]['sensor_data'])
                                    totallen += 1
                                
                                    break
  

                        if idx == int(intervalday)-1:
                            avg_data = round(totaldata/totallen, 4)
                            data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name, 'sensor_data': str(avg_data)})


        elif intervalday !='1' and len(intervalday) !=0:
            freq = intervalday+'d'
            sortdate = []
            realdata = pd.date_range(start=start, end=end, freq=freq)
            realdata = realdata.sort_values(ascending = False)

            for i in range(len(realdata)):
                date_i = str(realdata[i])
                sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

            sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
   

            data = []
           
            for i in range(len(sensordatas)):
              
                currnetdate = str(sensordatas[i]['sensor_data_date'])
               
                totaldata = 0
                totallen = 0
                if currnetdate[0:10] in sortdate:
                   
                    currentdaatime = currnetdate[11:20]
                    # print("currentdaatime",currentdaatime, currnetdate)
                    dateindex = sortdateall.index(currnetdate[0:10])
                  
                    for idx in range(int(intervalday)):
                        currentIndex = dateindex+idx
                       
                        if currentIndex < len(sortdateall):
                          
                            print("currentIndex", currentIndex)
                            lastdate = sortdateall[currentIndex] +" "+ currentdaatime
                            # print("lastdate", lastdate)

                            for j in range(len(sensordatas)):
                                # print("ininin", sensordatas[j]['sensor_data_date'], lastdate)
                                if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                    
                                    totaldata += float(sensordatas[j]['sensor_data'])
                                    totallen += 1
                                    
                                    break

                        if idx == int(intervalday)-1:
                            avg_data = round(totaldata/totallen, 4)
                            data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})

         



            

        else:
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
        

        fx_len = 0
        sensor_fx_list = []
        sensor_fx_name = []

        
        # test = eval('sin(1837)')
        # print(test)
        if sensor_fx1 and len(sensor_fx1) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx1_data')
            sensor_fx_name.append(sensor_fx1_name)
            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data']=round(result1,4)
     
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx2_data')
            sensor_fx_name.append(sensor_fx2_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
              

                data[i]['sensor_fx2_data']=round(result1,4)
           

        if sensor_fx3 and len(sensor_fx3) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx3_data')
            sensor_fx_name.append(sensor_fx3_name)
            for i in range(len(data)):
               
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
             
                result1 = eval(current_fx)
                data[i]['sensor_fx3_data']=round(result1,4)
           

        if sensor_fx4 and len(sensor_fx4) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx4_data')
            sensor_fx_name.append(sensor_fx4_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx4_data']=round(result1,4)
                

        if sensor_fx5 and len(sensor_fx5) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx5_data')
            sensor_fx_name.append(sensor_fx5_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx5_data']=round(result1,4)
        # rows = []       
        # for i in range(len(data)):
        #     if i == 0:
        #         rows.append("센서명","측정일시", "측정치")
        # print(pd.DataFrame([data]))
            
        if len(data) >= 0:
            sheet = book.create_sheet(sensor_name)
      
            sheet.cell(row=1, column=1).value = "센서명"
            sheet.cell(row=1, column=2).value = "측정일시"
            sheet.cell(row=1, column=3).value = "측정치"

            for i in range(fx_len):
                sheet.cell(row=1, column=3+i+1).value = sensor_fx_name[i]
            # sheet.cell(row=1, column=4).value = "변형각"
            # sheet.cell(row=1, column=5).value = "변화량1(MM)"
            # sheet.cell(row=1, column=6).value = "변화량1(MM)"
            # sheet.cell(row=1, column=7).value = "변화량1(MM)"
            # sheet.cell(row=1, column=8).value = "변화량1(MM)"
            # sheet.cell(row=1, column=8).value = "r"
            # sheet.cell(row=1, column=9).value = "WGS84 Z"

            # sheet.cell(row=1, column=1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=4).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=5).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=6).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=7).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=8).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        

            for idx in range(len(data)):


                sheet.cell(row=idx+2, column=1).value = sensor_display_name
                sheet.cell(row=idx+2, column=2).value = data[idx]['sensor_data_date']
                sheet.cell(row=idx+2, column=3).value = data[idx]['sensor_data']
                for i in range(fx_len):
                    sheet.cell(row=idx+2, column=3+i+1).value = data[idx][sensor_fx_list[i]]
                    
                # sheet.cell(row=idx+2, column=4).value = sensordata[idx]['변형각']
                # sheet.cell(row=idx+2, column=5).value = sensordata[idx]['변화량1(MM)']
                # sheet.cell(row=idx+2, column=6).value = sensordata[idx]['변화량2(MM)']
                # sheet.cell(row=idx+2, column=7).value = sensordata[idx]['변화량3(MM)']
                # sheet.cell(row=idx+2, column=8).value = sensordata[idx]['변화량4(MM)']
          
       
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        # print(pd.DataFrame(data))
        book.save(filename = excel_file + file_name)
       
        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')

        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        # wb = openpyxl.load_workbook("./static/contents/excel/"+file_name)
        # ws = wb[sensor_name]

    #  /Users/heeyeon/Documents/git/SJGEOTEC/static/contents/excel/2P03-2022-04-06_09:31:42.xlsx
    
        # li_chart = LineChart()
        # li_chart.title = 'test'
        # data_li = Reference(ws, min_col = 2, max_col = len(data), min_row = 1, max_row = len(data) )
        # print("1")
        # cats_li = Reference(ws, min_col = 1, max_col = 2, min_row = 4, max_row = fx_len+1 )
        # print("2")
        # li_chart.add_data(data_li,titles_from_data = True)
        # print("3")
        # li_chart.set_categories(cats_li)
        # print("4")
        # ws.add_chart(li_chart, "B2")
        # print("5")
        # wb.save("./static/contents/excel/"+file_name)
        # print("6")
        # print(data)

        return { "url" : filename }



class exceldownAll(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensor_name', type=str)

    def get(self):

        print("들어옴!!!!!!!!!!!!!!!!")

        sensorgroup_id =              request.args.get('sensorgroup_id')
    
        date_time_start =              request.args.get('date_time_start')
        date_time_end =              request.args.get('date_time_end')
        intervalday =              request.args.get('intervalday')
        time =              request.args.get('time')

        sensorgroup_type =              request.args.get('sensorgroup_type')

        sensordata_list = [dict(r) for r in SensorModel.find_by_datarogger_id_snesorgroup(sensorgroup_id)]


        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_interval = sensorgroup_obj.sensorgroup_interval


        totaldatalist = []
        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_initial_date = sensorgroup_obj.sensorgroup_initial_date

        
        sensor_gauge_factor = sensorgroup_obj.sensorgroup_initial_date
        for i in range(len(sensordata_list)):
            # print(sensordata_list[i]['sensor_name'])

            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']

            if sensor_initial_data is None or len(sensor_initial_data) == 0:
                sensor_initial_data = '0'
            if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
                sensor_gauge_factor = '0'

      

            sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(str(sensor_id))]

            

            
            if sensorlist:
                sensor_fx1 = sensorlist[0]['sensor_fx1']
                sensor_fx2 = sensorlist[0]['sensor_fx2']
                sensor_fx3 = sensorlist[0]['sensor_fx3']
                sensor_fx4 = sensorlist[0]['sensor_fx4']
                sensor_fx5 = sensorlist[0]['sensor_fx5']
                sensor_fx_check = sensorlist[0]['sensor_fx_check']


                # sensor_initial_data = sensorlist[0]['sensor_initial_data']

                sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
                sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
                sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
                sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
                sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
          

                format = "%Y-%m-%d"
                start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
                end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")
        

                dates = []
                dates = pd.date_range(start=start, end=end, freq='1d')
                dates = dates.sort_values(ascending = False)
                sortdateall = []
                currentIndex = 0

                for i in range(len(dates)):
                    date_i = str(dates[i])
                    sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])
                    
            
                param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

                if  len(time) ==0 and len(intervalday) !=0:
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)
                    # print(realdata)
                    # for i in range(len(realdata)):
                    #     date_i = str(realdata[i])
                    #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1
         
                    sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
                    
                    data = []   
                    for i in range(len(sensordatas)):
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate in sortdate:
                            # currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx

                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] 
                                
                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                        
                                            break
        

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name, 'sensor_data': str(avg_data)})
                
                elif intervalday !='1' and len(intervalday) !=0:
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1

                    sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
        

                    data = []
                
                    for i in range(len(sensordatas)):
                    
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate[0:10] in sortdate:
                        
                            currentdaatime = currnetdate[11:20]
                            # print("currentdaatime",currentdaatime, currnetdate)
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx
                            
                                if currentIndex < len(sortdateall):
                                
                                    print("currentIndex", currentIndex)
                                    lastdate = sortdateall[currentIndex] +" "+ currentdaatime
                                    # print("lastdate", lastdate)

                                    for j in range(len(sensordatas)):
                                        # print("ininin", sensordatas[j]['sensor_data_date'], lastdate)
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                            
                                            break

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})



                    

                else:
                    data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                if sensor_fx_check == '1' and sensor_fx1 and len(sensor_fx1) != 0 :
                    
                    for i in range(len(data)):
                        # print(data[i]['sensor_data'])
                        current_fx = sensor_fx1
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)
                    

                        data[i]['sensor_fx_data']=str(round(result1,4))
                   
                
                elif sensor_fx_check == '2' and sensor_fx2 and len(sensor_fx2) != 0 :
                   
                    for i in range(len(data)):
                        current_fx = sensor_fx2
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)

                        data[i]['sensor_fx_data']=str(round(result1,4))
            

                elif sensor_fx_check == '3' and sensor_fx3 and len(sensor_fx3) != 0 :
                  
                    for i in range(len(data)):
                        current_fx = sensor_fx3
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)

                        data[i]['sensor_fx_data']=str(round(result1,4))
         
                elif sensor_fx_check == '4' and sensor_fx4 and len(sensor_fx4) != 0 :
                    
                    for i in range(len(data)):
                        current_fx = sensor_fx4
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)

                        data[i]['sensor_fx_data']=str(round(result1,4))
                    
              

                elif sensor_fx_check == '5' and sensor_fx5 and len(sensor_fx5) != 0 :
                
                    for i in range(len(data)):
                    
                        current_fx = sensor_fx5
                    
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)
                    

                        data[i]['sensor_fx_data']=str(round(result1,4))
              

                # print(data)
                
                totaldatalist.append(data)
                
  

        sensorM = []

        if(sensorgroup_type == '0202'):
            sensorM.append('0m')

        for i in range(len(sensordata_list)):
            if sensorgroup_type == '0205':
                sensorM.append(sensordata_list[i]['sensor_display_name'])
            else:
                current_inteval = int(sensorgroup_interval) * (i+1)
                sensorM.append(str(current_inteval)+'m')

        print("sensorM", sensorM, sensorgroup_type)

        alldata = []

        if sensorgroup_type =='0201' or sensorgroup_type == '0205':
        # 가로형
            for idx in range(len(totaldatalist[0])):
                # print(totaldatalist[0][idx]['sensor_data_date'])
                alldata.append({'sensor_data_date': totaldatalist[0][idx]['sensor_data_date']})

                totaldata = 0
                for j in range(len(totaldatalist)):
                    current_idx = len(totaldatalist) - 1 -j
                    # print(current_idx, totaldatalist[current_idx][idx]['sensor_fx_data'])
                    # print(totaldata, totaldatalist[current_idx][idx]['sensor_fx_data'])
                    totaldata = float(totaldatalist[current_idx][idx]['sensor_fx_data'])
                    alldata[idx][sensorM[current_idx]] = str(round(totaldata, 4))

        elif sensorgroup_type =='0202':
            for idx in range(len(totaldatalist[0])):
                # print(totaldatalist[0][idx]['sensor_data_date'])
                alldata.append({'sensor_data_date': totaldatalist[0][idx]['sensor_data_date']})

                totaldata = 0
                for j in range(len(sensorM)):
                    current_idx = len(sensorM) - 1 -j
                    if j == 0 :
                        alldata[idx][sensorM[current_idx]] = '0'

                    else :

                        # print(current_idx, totaldatalist[current_idx][idx]['sensor_fx_data'])
                        # print(totaldata, totaldatalist[current_idx][idx]['sensor_fx_data'])
                        totaldata += float(totaldatalist[current_idx][idx]['sensor_fx_data'])
                        alldata[idx][sensorM[current_idx]] = str(round(totaldata, 4))


        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_name = sensorgroup_obj.sensorgroup_name

    
     
        

        # print(param)
        now = datetime.now()
        book = Workbook()

        file_name = sensorgroup_name +'-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

        row_len = 1
        col_len = 1
            
        if len(alldata) >= 0:
            sheet = book.create_sheet(sensorgroup_name)
      
     
            sheet.cell(row=1, column=1).value = "측정일시"
            for i in range(len(sensorM)):

                sheet.cell(row=1, column=i+2).value = sensorM[i]
    
            # sheet.cell(row=1, column=8).value = "r"
            # sheet.cell(row=1, column=9).value = "WGS84 Z"

            row_len += 1

            for idx in range(len(alldata)):

                sheet.cell(row=idx+2, column=1).value = alldata[idx]['sensor_data_date']

                for i in range(len(sensorM)):
                    sheet.cell(row=idx+2, column=2+i).value = alldata[idx][sensorM[i]]

                # if idx == 0:
                #     col_len += 1
                #     sheet.cell(row=row_len, column=1).value = alldata[idx]['sensor_data_date']
                #     sheet.cell(row=row_len, column=col_len).value = alldata[idx]['sensor_data']

                # else:

                #     for


                #     if sensordata[idx]['sensor_data_date'] == sensordata[idx-1]['sensor_data_date']:
              
                #         col_len += 1
        
                #         sheet.cell(row=row_len, column=1).value = sensordata[idx]['sensor_data_date']
                #         print(0,col_len, sensordata[idx]['sensor_name'], sensordata[col_len]['sensor_name'])
                #         if sensordata[idx]['sensor_name'] == sensordata[col_len-2]['sensor_name']:
                #             sheet.cell(row=row_len, column=col_len).value = sensordata[idx]['sensor_data']

                            
                #         # sheet.cell(row=idx+2, column=1).value = sensordata[idx]['sensor_name']
                #         # sheet.cell(row=idx+2, column=2).value = sensordata[idx]['sensor_data_date']
                #     else:
                #         row_len += 1
                #         col_len = 2

                #         sheet.cell(row=row_len, column=1).value = sensordata[idx]['sensor_data_date']
                        
                #         print(1,col_len, sensordata[idx]['sensor_name'], sensordata[col_len]['sensor_name'])
                #         if sensordata[idx]['sensor_name'] == sensordata[col_len-2]['sensor_name']:
                #             sheet.cell(row=row_len, column=col_len).value = sensordata[idx]['sensor_data']
                       

          

        # print(data)
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        print(excel_file+file_name)
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }


    def post(self):
        print("들어옴!!!!!!!!!!!!!!!!")

        params = exceldownAll.parse.parse_args()

        sensorgroup_id =              params['sensorgroup_id']
    
        date_time_start =              params['date_time_start']
        date_time_end =              params['date_time_end']
        intervalday =              params['intervalday']
        time =              params['time']

        sensordata_list = [dict(r) for r in SensorModel.find_by_datarogger_id_snesorgroup(sensorgroup_id)]
        datarogger_id =SensorModel.find_by_id(sensordata_list[0]['sensor_id']).datarogger_id

        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_name = sensorgroup_obj.sensorgroup_name
        sensorgroup_type = sensorgroup_obj.sensorgroup_type
        sensorgroup_initial_date = sensorgroup_obj.sensorgroup_initial_date

        sensor_gauge_factor = sensorgroup_obj.sensorgroup_initial_date


        sensordataname_list = ""
        for i in range(len(sensordata_list)):
            if(len(sensordataname_list) == 0):
                sensordataname_list += "'"+sensordata_list[i]['sensor_name']+"'"
            else:
                sensordataname_list += ", '"+sensordata_list[i]['sensor_name']+"'"


        totaldatalist = []
    

        


        datelist = []
        for i in range(len(sensordata_list)):
            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']

            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            if len(datelist) < len(data):
                for i in range(len(data)):
                    datelist.append({"sensor_data_date":data[i]['sensor_data_date']})
        
      

        for i in range(len(sensordata_list)):
            # print(sensordata_list[i]['sensor_name'])

            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']

         
            if sensor_initial_data is None or len(sensor_initial_data) == 0:
                sensor_initial_data = '0'
            if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
                sensor_gauge_factor = '0'

      

            sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(str(sensor_id))]

            # print("sensorlist", sensorlist)

            if sensorlist:
                sensor_fx1 = sensorlist[0]['sensor_fx1']
                sensor_fx2 = sensorlist[0]['sensor_fx2']
                sensor_fx3 = sensorlist[0]['sensor_fx3']
                sensor_fx4 = sensorlist[0]['sensor_fx4']
                sensor_fx5 = sensorlist[0]['sensor_fx5']
                sensor_fx_check = sensorlist[0]['sensor_fx_check']


                # sensor_initial_data = sensorlist[0]['sensor_initial_data']

                sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
                sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
                sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
                sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
                sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
          

                param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

     
                format = "%Y-%m-%d"
                start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
                end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")
        

                dates = []
                dates = pd.date_range(start=start, end=end, freq='1d')
                dates = dates.sort_values(ascending = False)
                sortdateall = []

                for i in range(len(dates)):
                    date_i = str(dates[i])
                    sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])


                if  len(time) ==0 and len(intervalday) !=0:
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)
                    # print(realdata)
                    # for i in range(len(realdata)):
                    #     date_i = str(realdata[i])
                    #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1
                    
                
                    sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
                    
                    data = []   
                    for i in range(len(sensordatas)):
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate in sortdate:
                            # currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx

                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] 
                                
                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                        
                                            break
        

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name, 'sensor_data': str(avg_data)})

  
                if len(intervalday) !=0 and int(intervalday) >1:
                    realdata = []
                    currentIndex=0
                    freq = intervalday+'d'
                    sortdate = []
                    # realdata = pd.date_range(start=start, end=end, freq=freq)
                    # realdata = realdata.sort_values(ascending = False)

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1

                    sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
   

                    data = []

                    for i in range(len(sensordatas)):
              
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate[0:10] in sortdate:
                         
                            currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                           
                            for idx in range(int(intervalday)):
                            
                                currentIndex = dateindex+idx
                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] +" "+ currentdaatime

                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                         
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                       
                                            break
                        
                                            


                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                  
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})

                 

                else:
               
                    data01 = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                    datalist01 = copy.deepcopy(datelist)
                
                    for i in range(len(datalist01)):
                        datayn = 0
                        for j in range(len(data01)):
                            
                            if datalist01[i]['sensor_data_date'] == data01[j]['sensor_data_date']:
                                datalist01[i]['sensor_name'] = data01[j]['sensor_name']
                                datalist01[i]['sensor_data'] = data01[j]['sensor_data']
                                datalist01[i]['sensor_fx_check_data'] = None
                                datayn =1
                                break

                        if datayn == 0:
                            datalist01[i]['sensor_name'] = sensor_name
                            datalist01[i]['sensor_data'] = None
                            datalist01[i]['sensor_fx_check_data'] = None


                    data = copy.deepcopy(datalist01)

     

                if sensor_fx1 and len(sensor_fx1) != 0 :
                  
                    for i in range(len(data)):
                      
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx1
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)
                        

                            data[i]['sensor_fx1_data']=str(round(result1,4))

                            if sensor_fx_check == '1':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))

                        else:
                           
                            data[i]['sensor_fx1_data']=None

            
                
                if sensor_fx2 and len(sensor_fx2) != 0 :
                   
                    for i in range(len(data)):

                        if data[i]['sensor_data']:
                            current_fx = sensor_fx2
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx2_data']=str(round(result1,4))

                            if sensor_fx_check == '2':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))

                        else:
                            data[i]['sensor_fx2_data']=None
           
                if sensor_fx3 and len(sensor_fx3) != 0 :
                  
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx3
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx3_data']=str(round(result1,4))

                            if sensor_fx_check == '3':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))
                        else:
                            data[i]['sensor_fx3_data']=None
                        
              

                if sensor_fx4 and len(sensor_fx4) != 0 :
                    
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx4
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx4_data']=str(round(result1,4))

                            if sensor_fx_check == '4':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))
                        else:

                            data[i]['sensor_fx4_data']=None
              

                if sensor_fx5 and len(sensor_fx5) != 0 :
                
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                    
                            current_fx = sensor_fx5
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)
                        

                            data[i]['sensor_fx5_data']=str(round(result1,4))

                            if sensor_fx_check == '5':
                                data[i]['sensor_fx_check_data']=str(round(result1,4))

                        else:
                            data[i]['sensor_fx5_data']=None
              

                totaldatalist.append(data)

  
        now = datetime.now()
        book = Workbook()


        file_name = sensorgroup_name +'-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

        row_len = 1
        col_len = 1
            
        if len(totaldatalist) >= 0 and len(totaldatalist[0]) > 0:
            sheet = book.create_sheet(sensorgroup_name)
      
     
            sheet.cell(row=1, column=1).value = "측정일시"
            for i in range(len(totaldatalist)):
                if sensorgroup_type == '0203':
                    sheet.cell(row=1, column=i+2).value = sensordata_list[i]['sensor_display_name']+"("+sensordata_list[i]['sensor_type']+")"
                else:
                    sheet.cell(row=1, column=i+2).value = sensordata_list[i]['sensor_display_name']

           

            row_len += 1

            for idx in range(len(totaldatalist[0])):
                sheet.cell(row=2+idx, column=1).value = totaldatalist[0][idx]['sensor_data_date']
                for j in range(len(totaldatalist)):
                    sheet.cell(row=2+idx, column=2+j).value = totaldatalist[j][idx]['sensor_fx_check_data']


                # if idx == 0:
                #     col_len += 1
                #     sheet.cell(row=row_len, column=1).value = totaldatalist[0][idx]['sensor_data_date']

                #     for j in range(len(totaldatalist)):
                #         sheet.cell(row=row_len, column=col_len).value = totaldatalist[0][idx]['sensor_fx_check_data']

                # else:


                #     if sensordata[idx]['sensor_data_date'] == sensordata[idx-1]['sensor_data_date']:
              
                #         col_len += 1
        
                #         sheet.cell(row=row_len, column=1).value = sensordata[idx]['sensor_data_date']
                      
                #         if sensordata[idx]['sensor_name'] == sensordata[col_len-2]['sensor_name']:
                #             sheet.cell(row=row_len, column=col_len).value = sensordata[idx]['sensor_data']

                            
                #         # sheet.cell(row=idx+2, column=1).value = sensordata[idx]['sensor_name']
                #         # sheet.cell(row=idx+2, column=2).value = sensordata[idx]['sensor_data_date']
                #     else:
                #         row_len += 1
                #         col_len = 2

                #         sheet.cell(row=row_len, column=1).value = sensordata[idx]['sensor_data_date']
                        
                       
                #         if sensordata[idx]['sensor_name'] == sensordata[col_len-2]['sensor_name']:
                #             sheet.cell(row=row_len, column=col_len).value = sensordata[idx]['sensor_data']
                       

   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')

        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }




class exceldownRecord(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_id', type=str)

    def get(self):

        print("들어옴!!!!!!!!!!!!!!!!")

 

        sensor_id =              request.args.get('sensor_id')
        date_time_start =              request.args.get('date_time_start')
        date_time_end =              request.args.get('date_time_end')
        
        sensor_obj = SensorModel.find_by_id(sensor_id)

        sensor_name = sensor_obj.sensor_name
        datarogger_id = sensor_obj.datarogger_id
        sensor_display_name = sensor_obj.sensor_display_name
     
        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_id)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']
        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'

     
        param = (sensor_name, str(datarogger_id),date_time_start,date_time_end)

        # print(param)
        now = datetime.now()
        book = Workbook()

        # print(sensordata)
        file_name = sensor_display_name +'-record-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

            
        data = [dict(r) for r in EditdataModel.sensor_data_all(param)]

        fx_len = 0
        sensor_fx_list = []
        sensor_fx_name = []

        if sensor_fx1 and len(sensor_fx1) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx1_data')
            sensor_fx_name.append(sensor_fx1_name)
            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data']=round(result1,4)
     
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx2_data')
            sensor_fx_name.append(sensor_fx2_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
              

                data[i]['sensor_fx2_data']=round(result1,4)
              

        if sensor_fx3 and len(sensor_fx3) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx3_data')
            sensor_fx_name.append(sensor_fx3_name)
            for i in range(len(data)):
               
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
             
                result1 = eval(current_fx)
                data[i]['sensor_fx3_data']=round(result1,4)
           

        if sensor_fx4 and len(sensor_fx4) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx4_data')
            sensor_fx_name.append(sensor_fx4_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx4_data']=round(result1,4)
                

        if sensor_fx5 and len(sensor_fx5) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx5_data')
            sensor_fx_name.append(sensor_fx5_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx5_data']=round(result1,4)
      
            
        if len(data) >= 0:
            sheet = book.create_sheet(sensor_name)
      
            sheet.cell(row=1, column=1).value = "센서명"
            sheet.cell(row=1, column=2).value = "측정일시"
            sheet.cell(row=1, column=3).value = "측정치"

            for i in range(fx_len):
                sheet.cell(row=1, column=3+i+1).value = sensor_fx_name[i]
 
        

            for idx in range(len(data)):


                sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_name']
                sheet.cell(row=idx+2, column=2).value = data[idx]['sensor_data_date']
                sheet.cell(row=idx+2, column=3).value = data[idx]['sensor_data']
                for i in range(fx_len):
                    sheet.cell(row=idx+2, column=3+i+1).value = data[idx][sensor_fx_list[i]]
                    
            
       
          




        # print(data)
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
      
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }





class excelupload(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_id', type=str)
    parse.add_argument('sensor_idx', type=str)
    parse.add_argument('sensor_idy', type=str)
    parse.add_argument('file', type=str)

    def post(self):
        params = excelupload.parse.parse_args()

        print("들어옴!!!!!!!!!!!!!!!!")

 
        file = request.files['file']
       
      
        sensor_id =              params['sensor_id']

        sensor_obj = SensorModel.find_by_id(sensor_id)
        sensor_name = sensor_obj.sensor_name
        datarogger_id = sensor_obj.datarogger_id

        datarogger_obj = DataroggerModel.find_by_id(datarogger_id)

        project_id = datarogger_obj.project_id      
        user_id = current_user.user_id
        create_date = datetime.now()
        modify_date = datetime.now()
     
        xlsx = pd.read_excel(file)
   
        col =xlsx.columns
        print(col)

        try:

            if sensor_id:
                for i in range(len(xlsx)):
                    sensor_data_date = xlsx[col[1]][i]
                    sensor_data = xlsx[col[2]][i]
                    # sensor_data_date_2 = sensor_data_date[0:4]+"."+sensor_data_date[5:7]+"."+sensor_data_date[8:]
                    # print(sensor_data_date, sensor_data, sensor_data_date_2)

                    sensor_X_index = sensor_obj.sensor_index


                    find_sensor_data = [dict(r) for r in EditdataModel.find_sensor_data_excel(sensor_name, str(datarogger_id), str(sensor_data_date))]
       

                    if len(find_sensor_data) !=0:
                        editdata_id = find_sensor_data[0]['editdata_id']
                        editdata_obj = EditdataModel.find_by_id(editdata_id)

                        if editdata_obj.use_yn == 'N':
                            sensor_data_list = [dict(r) for r in EditdataModel.find_sensor_edit_data_list(str(datarogger_id), str(sensor_data_date))]
                            for idx in range(len(sensor_data_list)):
                                editdata_id = sensor_data_list[idx]['editdata_id']

                                sensor_editdata_obj = EditdataModel.find_by_id(editdata_id)
                                sensor_editdata_obj.use_yn = 'Y'
                                sensor_editdata_obj.save_to_db()

                            

                        editdata_obj.sensor_data = sensor_data
                        editdata_obj.edit_yn = 'Y'
                        editdata_obj.use_yn = 'Y'
                        editdata_obj.modify_date = modify_date
                        editdata_obj.user_id = user_id
                        editdata_obj.save_to_db()

                        sensor_name_1 = editdata_obj.sensor_name
                        datarogger_id_1 = editdata_obj.datarogger_id

                        current_sensor_obj = SensorModel.find_by_datarogger_id_sensorname(sensor_name_1, datarogger_id_1)

                        if current_sensor_obj is not None:
                            current_sensor_id = current_sensor_obj.sensor_id

                            print("!")
                            alarm_obj = AlarmModel.find_by_sensor_id_datetime(datarogger_id, current_sensor_id, str(sensor_data_date))
                            if alarm_obj is not None:
                                print("INI:NININI", editdata_id)
                                print(alarm_obj.alarm_id)
                                alarm_obj.use_yn = 'N'
                                alarm_obj.alarm_status = 'Y'
                                alarm_obj.save_to_db()

                    else:
                        editdata_obj = EditdataModel(str(sensor_data_date), sensor_name, sensor_data, sensor_X_index,'Y','Y', user_id,datarogger_id,create_date,modify_date)
                        editdata_obj.save_to_db()

                message = "project_id: "+str(project_id)+ " sensor_id: "+sensor_id+ " sensor_name: " +sensor_name +" datarogger_id: " +str(datarogger_id) +" 데이터를 업로드 하였습니다."
                LogMessage.sensor_edit("data_edit", message, "0503")

            else:
                print("?")

            

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

       
        

        return { "url" : "filename" , "resultString":"excel upload 되었습니다."}


    def put(self):
        params = excelupload.parse.parse_args()
        print(params)
    

        print("put 들어옴!!!!!!!!!!!!!!!!")

 
        file = request.files['file']
      
        sensor_idx =              params['sensor_idx']
        sensor_idy =              params['sensor_idy']

        sensor_obj_x = SensorModel.find_by_id(sensor_idx)
        sensor_obj_y = SensorModel.find_by_id(sensor_idy)

        sensor_name_x = sensor_obj_x.sensor_name
        sensor_name_y = sensor_obj_y.sensor_name

        datarogger_id_x = sensor_obj_x.datarogger_id
        datarogger_id_y = sensor_obj_y.datarogger_id

        datarogger_obj = DataroggerModel.find_by_id(datarogger_id_y)
        project_id = datarogger_obj.project_id

        user_id = current_user.user_id
        create_date = datetime.now()
        modify_date = datetime.now()
     
        xlsx = pd.read_excel(file)

       
        # print(xlsx[0][1])
 
        col =xlsx.columns
      
        try:
            for i in range(len(xlsx)):
                sensor_data_date = xlsx[col[0]][i]
                sensor_data_x = xlsx[col[1]][i]
                sensor_data_y = xlsx[col[2]][i]

                sensor_X_index = sensor_obj_x.sensor_index
                sensor_Y_index = sensor_obj_y.sensor_index


                find_sensor_data_x = [dict(r) for r in EditdataModel.find_sensor_data_excel(sensor_name_x, str(datarogger_id_x), str(sensor_data_date))]
                find_sensor_data_y = [dict(r) for r in EditdataModel.find_sensor_data_excel(sensor_name_y, str(datarogger_id_y), str(sensor_data_date))]

                if len(find_sensor_data_x) !=0:
                    editdata_id_x = find_sensor_data_x[0]['editdata_id']
                    editdata_obj_x = EditdataModel.find_by_id(editdata_id_x)

                    if editdata_obj_x.use_yn == 'N':
                            sensor_data_list = [dict(r) for r in EditdataModel.find_sensor_edit_data_list(str(datarogger_id_x), str(sensor_data_date))]
                            for idx in range(len(sensor_data_list)):
                                editdata_id = sensor_data_list[idx]['editdata_id']

                                sensor_editdata_obj = EditdataModel.find_by_id(editdata_id)
                                sensor_editdata_obj.use_yn = 'Y'
                                sensor_editdata_obj.save_to_db()


                    editdata_obj_x.sensor_data = sensor_data_x
                    editdata_obj_x.user_id = user_id
                    editdata_obj_x.modify_date = modify_date
                    editdata_obj_x.edit_yn = 'Y'
                    editdata_obj_x.use_yn = 'Y'
                    editdata_obj_x.save_to_db()

                else:
                    editdata_obj_x = EditdataModel(str(sensor_data_date), sensor_name_x, sensor_data_x, sensor_X_index,'Y','Y', user_id,datarogger_id_x,create_date,modify_date)
                    editdata_obj_x.save_to_db()


                if len(find_sensor_data_y) !=0:
                  
                    editdata_id_y = find_sensor_data_y[0]['editdata_id']
                   
                    editdata_obj_y = EditdataModel.find_by_id(editdata_id_y)

                    if editdata_obj_y.use_yn == 'N':
                            sensor_data_list = [dict(r) for r in EditdataModel.find_sensor_edit_data_list(str(editdata_obj_y), str(sensor_data_date))]
                            for idx in range(len(sensor_data_list)):
                                editdata_id = sensor_data_list[idx]['editdata_id']

                                sensor_editdata_obj = EditdataModel.find_by_id(editdata_id)
                                sensor_editdata_obj.use_yn = 'Y'
                                sensor_editdata_obj.save_to_db()

                    editdata_obj_y.sensor_data = sensor_data_y
                    editdata_obj_y.user_id = user_id
                    editdata_obj_y.modify_date = modify_date
                    editdata_obj_y.edit_yn = 'Y'
                    editdata_obj_y.use_yn = 'Y'
                    editdata_obj_y.save_to_db()
                   

                else:
                    editdata_obj_y = EditdataModel(str(sensor_data_date), sensor_name_y, sensor_data_y, sensor_Y_index,'Y','Y', user_id,datarogger_id_y,create_date,modify_date)
                    editdata_obj_y.save_to_db()



            message = "project_id: "+str(project_id)+ " sensor_id: "+sensor_idx+ " sensor_name: " +sensor_name_x +" datarogger_id: " + str(datarogger_id_x) +" 데이터를 업로드 하였습니다."
            LogMessage.sensor_edit("data_edit", message, "0503")

            message = "project_id: "+str(project_id)+ " sensor_id: "+sensor_idy+ " sensor_name: " +sensor_name_y +" datarogger_id: " + str(datarogger_id_y) +" 데이터를 업로드 하였습니다."
            LogMessage.sensor_edit("data_edit", message, "0503")
         # 3개 등록하는 것   
        #     for i in range(len(xlsx)):
        #         # sensor_data_date = xlsx[i]['date']
        #         sensor_data_date = xlsx[col[0]][i]
        #         sensor_X = round(xlsx[col[1]][i],2)
        #         sensor_Y = round(xlsx[col[2]][i],2)
        #         sensor_C = round(xlsx[col[3]][i],2)
        #         # print(sensor_data_date, sensor_X, sensor_Y, sensor_C)

        #         sensor_X_obj = SensorModel.find_by_datarogger_id_sensorname( col[1], datarogger_id)
                
        #         sensor_X_index = sensor_X_obj.sensor_index
        #         print(sensor_data_date, col[1],sensor_X, sensor_X_index)
        #         editdata_X_obj = EditdataModel(sensor_data_date, col[1],sensor_X, sensor_X_index,'N', user_id,datarogger_id,create_date,modify_date)
        #         # editdata_X_obj.save_to_db()
                
        #         sensor_Y_obj = SensorModel.find_by_datarogger_id_sensorname( col[2], datarogger_id)
        #         sensor_Y_index = sensor_Y_obj.sensor_index
        #         print(sensor_data_date, col[2],sensor_Y, sensor_Y_index)
        #         editdata_Y_obj = EditdataModel(sensor_data_date, col[2],sensor_Y, sensor_Y_index,'N', user_id,datarogger_id,create_date,modify_date)
        #         # editdata_Y_obj.save_to_db()

        #         sensor_obj = SensorModel.find_by_datarogger_id_sensorname( col[3], datarogger_id)
        #         sensor_index = sensor_obj.sensor_index
        #         print(sensor_data_date, col[3],sensor_C, sensor_index)
        #         editdata_obj = EditdataModel(sensor_data_date, col[3],sensor_C, sensor_index,'N', user_id,datarogger_id,create_date,modify_date)
        #         # editdata_obj.save_to_db()

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

       
        

        return { "url" : "filename" , "resultString":"excel upload 되었습니다."}


class editdataroadcell(Resource):


    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('sensorgroup_type', type=str)

    def get(self):

        print("들어옴!!!!!!!!!!!!!!!!")

        sensorgroup_id =              request.args.get('sensorgroup_id')
    
        date_time_start =              request.args.get('date_time_start')
        date_time_end =              request.args.get('date_time_end')
        intervalday =              request.args.get('intervalday')
        time =              request.args.get('time')

        sensorgroup_type =              request.args.get('sensorgroup_type')

        sensordata_list = [dict(r) for r in SensorModel.find_by_datarogger_id_snesorgroup(sensorgroup_id)]

        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)

        sensorgroup_fx = sensorgroup_obj.sensorgroup_fx
        avgagugefactor = sensorgroup_obj.sensorgroup_gauge_factor
        function_name = FunctionModel.find_by_id(str(sensorgroup_obj.sensorgroup_fx)).function_name

    
        totalinitial = 0
        totalgaugefactor = 0
        # gaugelen = len(sensordata_list)
        roadcelldata = []

        datelist = []
        for i in range(len(sensordata_list)):
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']
            if sensor_gauge_factor is None:
                sensor_gauge_factor = '0.002'
            sensor_fx_check = sensordata_list[i]['sensor_fx_check']

            sensor_fx1 = sensordata_list[0]['sensor_fx1']
            sensor_fx2 = sensordata_list[0]['sensor_fx2']
            sensor_fx3 = sensordata_list[0]['sensor_fx3']
            sensor_fx4 = sensordata_list[0]['sensor_fx4']
            sensor_fx5 = sensordata_list[0]['sensor_fx5']
            
            # if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
            #     sensor_gauge_factor = 0
            #     gaugelen -= 1

            totalinitial += float(sensordata_list[i]['sensor_initial_data'])
            # totalgaugefactor += float(sensor_gauge_factor)

            sensor_id = sensordata_list[i]['sensor_id']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_name = sensordata_list[i]['sensor_name']
            sensor_obj = SensorModel.find_by_id(sensor_id)
           
        
            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            # roadcelldata.append(data) 

            if len(datelist) < len(data):
                for j in range(len(data)):
                    datelist.append({"sensor_data_date":data[j]['sensor_data_date']})


            start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
            end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")
            dates = []
            dates = pd.date_range(start=start, end=end, freq='1d')
            dates = dates.sort_values(ascending = False)
            sortdateall = []
            print(start, end, dates)
            for j in range(len(dates)):
                date_i = str(dates[j])
                sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

            print("sortdateall", sortdateall)

            if  len(time) ==0 and len(intervalday) !=0:
                freq = intervalday+'d'
                sortdate = []
                realdata = pd.date_range(start=start, end=end, freq=freq)
                realdata = realdata.sort_values(ascending = False)
                # print(realdata)
                # for i in range(len(realdata)):
                #     date_i = str(realdata[i])
                #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

                sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                befor_days = datetime.now()
                day_i = 1
                while start < befor_days:
                    befor_days = end - timedelta(days= day_i * int(intervalday))
                    befor_one_days = str(befor_days)
                    sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                    day_i += 1
                    print(befor_days)
            
                sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
                
                data = []   
                for i in range(len(sensordatas)):
                    currnetdate = str(sensordatas[i]['sensor_data_date'])
                
                    totaldata = 0
                    totallen = 0
                    if currnetdate in sortdate:
                        # currentdaatime = currnetdate[11:20]
                        dateindex = sortdateall.index(currnetdate[0:10])
                    
                        for idx in range(int(intervalday)):
                            currentIndex = dateindex+idx

                            if currentIndex < len(sortdateall):
                                lastdate = sortdateall[currentIndex] 
                            
                                for j in range(len(sensordatas)):
                                    if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                        totaldata += float(sensordatas[j]['sensor_data'])
                                        totallen += 1
                                    
                                        break
    

                            if idx == int(intervalday)-1:
                                avg_data = round(totaldata/totallen, 4)
                                data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name, 'sensor_data': str(avg_data)})





            elif len(intervalday) !=0 and int(intervalday) > 1:
                freq = intervalday+'d'
                sortdate = []
                sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                befor_days = datetime.now()
                day_i = 1
                while start < befor_days:
                    befor_days = end - timedelta(days= day_i * int(intervalday))
                    befor_one_days = str(befor_days)
                    sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                    day_i += 1

                sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
    

                data = []
            
                for i in range(len(sensordatas)):
                
                    currnetdate = str(sensordatas[i]['sensor_data_date'])
                
                    totaldata = 0
                    totallen = 0
                    if currnetdate[0:10] in sortdate:
                    
                        currentdaatime = currnetdate[11:20]
                        # print("currentdaatime",currentdaatime, currnetdate)
                        dateindex = sortdateall.index(currnetdate[0:10])
                    
                        for idx in range(int(intervalday)):
                            currentIndex = dateindex+idx
                        
                            if currentIndex < len(sortdateall):
                            
                                lastdate = sortdateall[currentIndex] +" "+ currentdaatime

                                for j in range(len(sensordatas)):
                                    if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                        
                                        totaldata += float(sensordatas[j]['sensor_data'])
                                        totallen += 1
                                        
                                        break

                            if idx == int(intervalday)-1:
                                avg_data = round(totaldata/totallen, 4)
                                data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})
                                        
            else:
 
                data01 = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                datalist01 = copy.deepcopy(datelist)
   
                for i in range(len(datalist01)):
                    datayn = 0
                    for j in range(len(data01)):
                        
                        if datalist01[i]['sensor_data_date'] == data01[j]['sensor_data_date']:
                            datalist01[i]['sensor_name'] = data01[j]['sensor_name']
                            datalist01[i]['sensor_data'] = data01[j]['sensor_data']
                            datayn =1
                            break

                    if datayn == 0:
                        datalist01[i]['sensor_name'] = sensor_name
                        datalist01[i]['sensor_data'] = None


                data = copy.deepcopy(datalist01)
                

            if sensor_fx_check == '1' and sensor_fx1 is not None and len(sensor_fx1) != 0 :
                    
                for i in range(len(data)):
                    if data[i]['sensor_data']:
                        current_fx = sensor_fx1
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)
                    

                        data[i]['sensor_fx_data']=str(round(result1,4))

                    else:
                        data[i]['sensor_fx_data']=None
            
            elif sensor_fx_check == '2' and sensor_fx2 is not None and len(sensor_fx2) != 0 :
                
                for i in range(len(data)):

                    if data[i]['sensor_data']:
                        current_fx = sensor_fx2
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)

                        data[i]['sensor_fx_data']=str(round(result1,4))

                    else:
                        data[i]['sensor_fx_data']=None
        

            elif sensor_fx_check == '3' and sensor_fx3 is not None and len(sensor_fx3) != 0 :
                
                for i in range(len(data)):
                    if data[i]['sensor_data']:
                        current_fx = sensor_fx3
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)

                        data[i]['sensor_fx_data']=str(round(result1,4))
                    else:
                        data[i]['sensor_fx_data']=None
        
            elif sensor_fx_check == '4' and sensor_fx4 is not None and len(sensor_fx4) != 0 :
                
                for i in range(len(data)):
                    if data[i]['sensor_data']:
                        current_fx = sensor_fx4
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)

                        data[i]['sensor_fx_data']=str(round(result1,4))
                    else:

                        data[i]['sensor_fx_data']=None
                
            

            elif sensor_fx_check == '5' and sensor_fx5 is not None and len(sensor_fx5) != 0 :
            
                for i in range(len(data)):
                    if data[i]['sensor_data']:
                
                        current_fx = sensor_fx5
                        current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                        current_fx = current_fx.replace('$ini', sensor_initial_data)
                        current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                        result1 = eval(current_fx)
                    

                        data[i]['sensor_fx_data']=str(round(result1,4))

                    else:
                        data[i]['sensor_fx_data']=None
            


            roadcelldata.append(data)

        avginitial = round(totalinitial / len(sensordata_list) , 4)
        # avgagugefactor = round(totalgaugefactor / len(sensordata_list) , 4)

        avglist = []
      

        for i in range(len(roadcelldata[0])):
            datalen = 0
            current_data = roadcelldata[0]
            avglist.append({'sensor_data_date': current_data[i]['sensor_data_date']})
            datavalue = 0
            for j in range(len(roadcelldata)):
                avglist[i][roadcelldata[j][i]['sensor_name']] = roadcelldata[j][i]['sensor_fx_data']
           
                if roadcelldata[j][i]['sensor_data']:
                    datavalue += float(roadcelldata[j][i]['sensor_data'])
                    datalen += 1

            avgdata = round(datavalue / datalen, 4)
            avglist[i]['avg_data'] = avgdata


        if sensorgroup_fx:
            function_obj = FunctionModel.find_by_id(str(sensorgroup_fx))

            fomula = function_obj.function_formula

            current_fx = fomula
                
            current_fx = current_fx.replace('$sen', str(avglist[i]['avg_data']))
            current_fx = current_fx.replace('$ini', str(avginitial))
            current_fx = current_fx.replace('$GF', str(avgagugefactor))

            initial_fx = round(eval(current_fx), 4)
 
        if sensorgroup_fx:
            function_obj = FunctionModel.find_by_id(str(sensorgroup_fx))

            fomula = function_obj.function_formula
         

            for i in range(len(avglist)):
                current_fx = fomula
                
                current_fx = current_fx.replace('$sen', str(avglist[i]['avg_data']))
                current_fx = current_fx.replace('$ini', str(avginitial))
                current_fx = current_fx.replace('$GF', str(avgagugefactor))
                
                result1 = round(eval(current_fx), 4)

                avglist[i]['sensor_fx_data'] = result1
                avglist[i]['initial_fx'] = initial_fx

        else : 

            for i in range(len(avglist)):
                avglist[i]['sensor_fx_data'] = avglist[i]['avg_data']
                avglist[i]['initial_fx'] = avginitial


        now = datetime.now()
        book = Workbook()
        # sensordata = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
        # print(sensordata)
        file_name = sensorgroup_obj.sensorgroup_name +'-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

        if len(avglist) >= 0:
            sheet = book.create_sheet(sensor_name)
      
            sheet.cell(row=1, column=1).value = "측정일시"
            for i in range(len(sensordata_list)):
                sheet.cell(row=1, column=i+2).value = sensordata_list[i]['sensor_display_name']
       
            sheet.cell(row=1, column=2+len(sensordata_list)).value = function_name


            for idx in range(len(avglist)):


                sheet.cell(row=idx+2, column=1).value = avglist[idx]['sensor_data_date']
                for i in range(len(sensordata_list)):
                    sheet.cell(row=idx+2, column=i+2).value = avglist[idx][sensordata_list[i]['sensor_name']]
                   
                sheet.cell(row=idx+2, column=2+len(sensordata_list)).value = avglist[idx]['sensor_fx_data']
        
              
                    
            
       
          




        # print(data)
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
      
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }

    def post(self):

        print("roadcell_data!!")
        params = editdatatableall.parse.parse_args()
       
        sensorgroup_id = params["sensorgroup_id"]
        date_time_start = params["date_time_start"]
        date_time_end = params["date_time_end"]
        intervalday = params["intervalday"]
        time = params["time"]
        sensorgroup_type = params["sensorgroup_type"]
    

        sensordata_list = [dict(r) for r in SensorModel.find_by_datarogger_id_snesorgroup(sensorgroup_id)]

        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)

        sensorgroup_fx = sensorgroup_obj.sensorgroup_fx
        avgagugefactor = sensorgroup_obj.sensorgroup_gauge_factor

       

        totalinitial = 0
        # totalgaugefactor = 0
        # gaugelen = len(sensordata_list)
        roadcelldata = []
        roadcelldata01 = []

        datelist = []
        for i in range(len(sensordata_list)):
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_name = sensordata_list[i]['sensor_name']
            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
            if len(datelist) < len(data):
                for i in range(len(data)):
                    datelist.append({"sensor_data_date":data[i]['sensor_data_date']})

        for i in range(len(sensordata_list)):
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            if sensor_initial_data is None :
                sensor_initial_data = 0
            # sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']
            
            # if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
            #     sensor_gauge_factor = 0
            #     gaugelen -= 1

            totalinitial += float(sensor_initial_data)
            # totalgaugefactor += float(sensor_gauge_factor)

            sensor_id = sensordata_list[i]['sensor_id']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_name = sensordata_list[i]['sensor_name']
            sensor_obj = SensorModel.find_by_id(sensor_id)
           
        
            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            # roadcelldata.append(data) 

            if len(datelist) < len(data):
                for i in range(len(data)):
                    datelist.append({"sensor_data_date":data[i]['sensor_data_date']})


            start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
            end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")

            dates = []
            dates = pd.date_range(start=start, end=end, freq='1d')
            dates = dates.sort_values(ascending = False)
            sortdateall = []
            currentIndex = 0

            
            for i in range(len(dates)):
                date_i = str(dates[i])
                sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])
                
        
            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

            if  len(time) ==0 and len(intervalday) !=0:
    
                sortdate = []
               
                sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                befor_days = datetime.now()
                day_i = 1
                while start < befor_days:
                    befor_days = end - timedelta(days= day_i * int(intervalday))
                    befor_one_days = str(befor_days)
                    sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                    day_i += 1
                
            
                sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
                
                data = []   
                for i in range(len(sensordatas)):
                    currnetdate = str(sensordatas[i]['sensor_data_date'])
                
                    totaldata = 0
                    totallen = 0
                    if currnetdate in sortdate:
                        # currentdaatime = currnetdate[11:20]
                        dateindex = sortdateall.index(currnetdate[0:10])
                    
                        for idx in range(int(intervalday)):
                            currentIndex = dateindex+idx

                            if currentIndex < len(sortdateall):
                                lastdate = sortdateall[currentIndex] 
                            
                                for j in range(len(sensordatas)):
                                    if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                        totaldata += float(sensordatas[j]['sensor_data'])
                                        totallen += 1
                                    
                                        break
    

                            if idx == int(intervalday)-1:
                                avg_data = round(totaldata/totallen, 4)
                                data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name, 'sensor_data': str(avg_data)})


            elif len(intervalday) !=0 and int(intervalday) > 1:
             
                sortdate = []
               

                sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                befor_days = datetime.now()
                day_i = 1
                while start < befor_days:
                    befor_days = end - timedelta(days= day_i * int(intervalday))
                    befor_one_days = str(befor_days)
                    sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                    day_i += 1
          

                sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
    

                data = []
                # for j in range(len(datelist)):
                #     data.append({'sensor_data_date': datelist[i]['sensor_data_date'], 'sensor_name':sensor_name, 'sensor_data':None})
                # print(data)
            
                for i in range(len(sensordatas)):
                
                    currnetdate = str(sensordatas[i]['sensor_data_date'])
                
                    totaldata = 0
                    totallen = 0
                    if currnetdate[0:10] in sortdate:
                    
                        currentdaatime = currnetdate[11:20]
                        dateindex = sortdateall.index(currnetdate[0:10])
                    
                        for idx in range(int(intervalday)):
                            currentIndex = dateindex+idx
                        
                            if currentIndex < len(sortdateall):
                            
                            
                                lastdate = sortdateall[currentIndex] +" "+ currentdaatime               

                                for j in range(len(sensordatas)):
                                    if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                        
                                        totaldata += float(sensordatas[j]['sensor_data'])
                                        totallen += 1
                                        
                                        break

                            if idx == int(intervalday)-1:
                                avg_data = round(totaldata/totallen, 4)
                                # for j in range(len(data)):
                                #     if data[j]['sensor_data_date'] == sensordatas[i]['sensor_data_date']:
                                #         data[j]['sensor_data'] = str(avg_data)
                                #         break
                                
                                data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})
                        
                        # roadcelldata.append(data)

                    
            else:

                data01 = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                datalist01 = copy.deepcopy(datelist)
                for i in range(len(datalist01)):
                    datayn = 0
                    for j in range(len(data01)):
                        
                        if datalist01[i]['sensor_data_date'] == data01[j]['sensor_data_date']:
                            datalist01[i]['sensor_name'] = data01[j]['sensor_name']
                            datalist01[i]['sensor_data'] = data01[j]['sensor_data']
                            datayn =1
                            break

                    if datayn == 0:
                        datalist01[i]['sensor_name'] = sensor_name
                        datalist01[i]['sensor_data'] = None


                data = copy.deepcopy(datalist01)
            roadcelldata01.append(data)

        avginitial = round(totalinitial / len(sensordata_list) , 4)
      
        avglist = []
        # for i in range(len(roadcelldata) ):
        #     print(len(roadcelldata[i]))

        print(len(roadcelldata01[0]))
        print(len(roadcelldata01[1]))
        print(len(roadcelldata01))
        newdatelist = []
        for i in range(len(roadcelldata01)):
            datayn = 0
            if len(newdatelist) < len(roadcelldata01[i]):
                newdatelist = []
                for j in range(len(roadcelldata01[i])):
                    newdatelist.append({'sensor_data_date': roadcelldata01[i][j]['sensor_data_date']})
        print("newdatelist",len(newdatelist))

        newdata = []
       

        for i in range(len(roadcelldata01)):
            newdata = []
            for j in range(len(newdatelist)):
                newdata.append({'sensor_data_date':newdatelist[j]['sensor_data_date'], 'sensor_name':roadcelldata01[i][0]['sensor_name'], 'sensor_data': None})

                for x in range(len(roadcelldata01[i])):
                    if newdatelist[j]['sensor_data_date'] == roadcelldata01[i][x]['sensor_data_date']:
                        newdata[j]['sensor_data'] = roadcelldata01[i][x]['sensor_data']
                        break
            print("newdata",len(newdata))
            
            roadcelldata.append(newdata)
                    
        print(len(roadcelldata[0]))
        print(len(roadcelldata[1]))
        print(len(roadcelldata))

        for i in range(len(roadcelldata[0])):
            datalen = 0
            current_data = roadcelldata[0]
            
            avglist.append({'sensor_data_date': current_data[i]['sensor_data_date']})
            datavalue = 0
            for j in range(len(roadcelldata)):
               
                avglist[i][roadcelldata[j][i]['sensor_name']] = roadcelldata[j][i]['sensor_data']
           
               
                if roadcelldata[j][i]['sensor_data']:
                    datavalue += float(roadcelldata[j][i]['sensor_data'])
                    datalen += 1

            avgdata = round(datavalue / datalen, 4)
            avglist[i]['avg_data'] = avgdata

               

        if len(avglist) == 0:
            return { "url" : "filename" , "resultString":"SUCCESS", "data":avglist}

        if sensorgroup_fx:
            function_obj = FunctionModel.find_by_id(str(sensorgroup_fx))

            fomula = function_obj.function_formula

            current_fx = fomula
                
            current_fx = current_fx.replace('$sen', str(avglist[i]['avg_data']))
            current_fx = current_fx.replace('$ini', str(avginitial))
            current_fx = current_fx.replace('$GF', str(avgagugefactor))


            initial_fx = round(eval(current_fx), 4)

         


 
        if sensorgroup_fx:
            function_obj = FunctionModel.find_by_id(str(sensorgroup_fx))

            fomula = function_obj.function_formula
         

            for i in range(len(avglist)):
                current_fx = fomula
                
                current_fx = current_fx.replace('$sen', str(avglist[i]['avg_data']))
                current_fx = current_fx.replace('$ini', str(avginitial))
                current_fx = current_fx.replace('$GF', str(avgagugefactor))
                
                result1 = round(eval(current_fx), 4)

                avglist[i]['sensor_fx_data'] = result1
                avglist[i]['initial_fx'] = initial_fx

        else : 

            for i in range(len(avglist)):
                avglist[i]['sensor_fx_data'] = avglist[i]['avg_data']
                avglist[i]['initial_fx'] = avginitial



        project_id = sensorgroup_obj.project_id

        fx_list = [dict(r) for r in FunctionModel.find_by_roadcell(str(project_id))]

        namelist = []
        for i in range(len(fx_list)):
            namelist.append(fx_list[i]['function_name'])
            for j in range(len(avglist)):
                current_fx = fx_list[i]['function_formula']
                
                current_fx = current_fx.replace('$sen', str(avglist[j]['avg_data']))
                current_fx = current_fx.replace('$ini', str(avginitial))
                current_fx = current_fx.replace('$GF', str(avgagugefactor))
                
                result1 = round(eval(current_fx), 4)

                avglist[j][fx_list[i]['function_name']] = result1
                avglist[j]['sensor_fx_len'] = len(fx_list)
              



        

        return { "url" : "filename" , "resultString":"SUCCESS", "data":avglist, "namelist":namelist}




class exceldownscatter(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_idx', type=str)
    parse.add_argument('sensor_idy', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensor_name', type=str)

    def get(self):

        print("!!!!들어옴!!!!!!!!!!!!!!!!")

       
        
        
        date_time_start =              request.args.get('date_time_start')
        date_time_end =              request.args.get('date_time_end')
        intervalday =              request.args.get('intervalday')
        time =              request.args.get('time')
        sensor_idx =              request.args.get('sensor_idx')
        sensor_idy =              request.args.get('sensor_idy')

    

        sensor_obj = SensorModel.find_by_id(sensor_idx)
        sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_idx)
        datarogger_id = sensor_obj.datarogger_id
        sensor_name = sensor_obj.sensor_name
        sensor_display_name = sensor_obj.sensor_display_name
     
       

        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_idx)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']

        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'


        

   
        now = datetime.now()
        book = Workbook()
        file_name = sensor_display_name +'-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

       
        start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
        end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")

        realdata = []
        currentIndex = 0

       
        dates = []
        dates = pd.date_range(start=start, end=end, freq='1d')
        dates = dates.sort_values(ascending = False)
        sortdateall = []

        for i in range(len(dates)):
            date_i = str(dates[i])
            sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])
            
    
        param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
        
        if intervalday !='1' and len(intervalday) !=0:
            freq = intervalday+'d'
            sortdate = []
            realdata = pd.date_range(start=start, end=end, freq=freq)
            realdata = realdata.sort_values(ascending = False)

            for i in range(len(realdata)):
                date_i = str(realdata[i])
                sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

            sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
   

            data = []
           
            for i in range(len(sensordatas)):
              
                currnetdate = str(sensordatas[i]['sensor_data_date'])
               
                totaldata = 0
                totallen = 0
                if currnetdate[0:10] in sortdate:
                   
                    currentdaatime = currnetdate[11:20]
                    dateindex = sortdateall.index(currnetdate[0:10])
                  
                    for idx in range(int(intervalday)):
                        currentIndex = dateindex+idx
                       
                        if currentIndex < len(sortdateall):
                          
                            print("currentIndex", currentIndex)
                            lastdate = sortdateall[currentIndex] +" "+ currentdaatime

                            for j in range(len(sensordatas)):
                                if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                    
                                    totaldata += float(sensordatas[j]['sensor_data'])
                                    totallen += 1
                                    
                                    break

                        if idx == int(intervalday)-1:
                            avg_data = round(totaldata/totallen, 4)
                            data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})
        else:
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

        # scatter y데이터 

        sensorlist_y = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_idy)]
     
        sensor_gauge_factor_y = sensorlist_y[0]['sensor_gauge_factor']
        sensor_initial_data_y = sensorlist_y[0]['sensor_initial_data']
        if sensor_initial_data_y is None:
            sensor_initial_data_y = '0'
        if sensor_gauge_factor_y is None:
            sensor_gauge_factor_y = '0'

        sensor_fx1_y = sensorlist_y[0]['sensor_fx1']
        sensor_fx2_y = sensorlist_y[0]['sensor_fx2']
        sensor_fx3_y = sensorlist_y[0]['sensor_fx3']
        sensor_fx4_y = sensorlist_y[0]['sensor_fx4']
        sensor_fx5_y = sensorlist_y[0]['sensor_fx5']

        sensor_fx1_name_y = sensorlist_y[0]['sensor_fx1_name']
        sensor_fx2_name_y = sensorlist_y[0]['sensor_fx2_name']
        sensor_fx3_name_y = sensorlist_y[0]['sensor_fx3_name']
        sensor_fx4_name_y = sensorlist_y[0]['sensor_fx4_name']
        sensor_fx5_name_y = sensorlist_y[0]['sensor_fx5_name']

        sensor_obj_y = SensorModel.find_by_id(sensor_idy)
        datarogger_id_y = sensor_obj_y.datarogger_id
        sensor_name_y = sensor_obj_y.sensor_name

        param_y= (sensor_name_y, str(datarogger_id_y), date_time_start, date_time_end, '', time )


        if intervalday !='1' and len(intervalday) !=0:
            freq = intervalday+'d'
            sortdate = []
            realdata = pd.date_range(start=start, end=end, freq=freq)
            realdata = realdata.sort_values(ascending = False)

            for i in range(len(realdata)):
                date_i = str(realdata[i])
                sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

            sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param_y)]
   

            data_y = []
           
            for i in range(len(sensordatas)):
              
                currnetdate = str(sensordatas[i]['sensor_data_date'])
               
                totaldata = 0
                totallen = 0
                if currnetdate[0:10] in sortdate:
                   
                    currentdaatime = currnetdate[11:20]
                    dateindex = sortdateall.index(currnetdate[0:10])
                  
                    for idx in range(int(intervalday)):
                        currentIndex = dateindex+idx
                       
                        if currentIndex < len(sortdateall):
                          
                            lastdate = sortdateall[currentIndex] +" "+ currentdaatime

                            for j in range(len(sensordatas)):
                                if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                    
                                    totaldata += float(sensordatas[j]['sensor_data'])
                                    totallen += 1
                                    
                                    break

                        if idx == int(intervalday)-1:
                            avg_data = round(totaldata/totallen, 4)
                            data_y.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})
        else:
            data_y = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param_y)]

            
        for idx in range(len(data_y)):
            data[idx]['sensor_data_y'] = data_y[idx]['sensor_data']

        fx_len = 0
        sensor_fx_list = []
        sensor_fx_name = []

      
        if sensor_fx1 and len(sensor_fx1) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx1_data')
            sensor_fx_list.append('sensor_fx1_data_y')
            sensor_fx_name.append(sensor_fx1_name+"(x)")
            sensor_fx_name.append(sensor_fx1_name+"(y)")
            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
     
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx2_data')
            sensor_fx_list.append('sensor_fx2_data_y')
            sensor_fx_name.append(sensor_fx2_name+"(x)")
            sensor_fx_name.append(sensor_fx2_name+"(y)")
            for i in range(len(data)):
                
                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
              

                data[i]['sensor_fx2_data']=round(result1,4)
                print(data[i])
            for i in range(len(data)):

                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx2_data_y']=round(result1,4)

        if sensor_fx3 and len(sensor_fx3) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx3_data')
            sensor_fx_list.append('sensor_fx3_data_y')
            sensor_fx_name.append(sensor_fx3_name+"(x)")
            sensor_fx_name.append(sensor_fx3_name+"(y)")
            for i in range(len(data)):
               
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
             
                result1 = eval(current_fx)
                data[i]['sensor_fx3_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx3_data_y']=round(result1,4)
           

        if sensor_fx4 and len(sensor_fx4) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx4_data')
            sensor_fx_list.append('sensor_fx4_data_y')
            sensor_fx_name.append(sensor_fx4_name+"(x)")
            sensor_fx_name.append(sensor_fx4_name+"(y)")
            for i in range(len(data)):
                
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx4_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx4_data_y']=round(result1,4)
                

        if sensor_fx5 and len(sensor_fx5) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx5_data')
            sensor_fx_list.append('sensor_fx5_data_y')
            sensor_fx_name.append(sensor_fx5_name+"(x)")
            sensor_fx_name.append(sensor_fx5_name+"(y)")
            for i in range(len(data)):
                
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx5_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx5_data_y']=round(result1,4)
     
        # rows = []       
        # for i in range(len(data)):
        #     if i == 0:
        #         rows.append("센서명","측정일시", "측정치")
        # print(pd.DataFrame([data]))
            
        if len(data) >= 0:
            sheet = book.create_sheet(sensor_name)
      
            # sheet.cell(row=1, column=1).value = "센서명"
            sheet.cell(row=1, column=1).value = "측정일시"
            sheet.cell(row=1, column=2).value = sensor_display_name+"(x)"
            sheet.cell(row=1, column=3).value = sensor_display_name+"(y)"

            for i in range(fx_len):
                sheet.cell(row=1, column=3+i+1).value = sensor_fx_name[i]
            # sheet.cell(row=1, column=4).value = "변형각"
            # sheet.cell(row=1, column=5).value = "변화량1(MM)"
            # sheet.cell(row=1, column=6).value = "변화량1(MM)"
            # sheet.cell(row=1, column=7).value = "변화량1(MM)"
            # sheet.cell(row=1, column=8).value = "변화량1(MM)"
            # sheet.cell(row=1, column=8).value = "r"
            # sheet.cell(row=1, column=9).value = "WGS84 Z"

            # sheet.cell(row=1, column=1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=4).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=5).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=6).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=7).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # sheet.cell(row=1, column=8).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        

            for idx in range(len(data)):


                # sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_name']
                sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_data_date']
                sheet.cell(row=idx+2, column=2).value = data[idx]['sensor_data']
                sheet.cell(row=idx+2, column=3).value = data[idx]['sensor_data_y']
                for i in range(fx_len):
                    sheet.cell(row=idx+2, column=3+i+1).value = data[idx][sensor_fx_list[i]]
     
          
       
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        # print(pd.DataFrame(data))
        book.save(filename = excel_file + file_name)
       
        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        print(excel_file+file_name)
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

 


        return { "url" : filename }



    def post(self):

        print("!!!!들어옴!!!!!!!!!!!!!!!!")
        params = exceldownscatter.parse.parse_args()
     
        sensor_idx =              params['sensor_idx']
        sensor_idy =              params['sensor_idy']
        date_time_start =              params['date_time_start']
        date_time_end =              params['date_time_end']

    
        sensor_obj = SensorModel.find_by_id(sensor_idx)
        sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_idx)
        datarogger_id = sensor_obj.datarogger_id
        sensor_name = sensor_obj.sensor_name
        sensor_display_name = sensor_obj.sensor_display_name
     
        param = (sensor_name, str(datarogger_id) ,date_time_start, date_time_end)

        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_idx)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']
        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'


        sensorlist_y = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_idy)]
        # sensor_fx1 = sensorlist[0]['sensor_fx1']
        # sensor_fx2 = sensorlist[0]['sensor_fx2']
        # sensor_fx3 = sensorlist[0]['sensor_fx3']
        # sensor_fx4 = sensorlist[0]['sensor_fx4']
        # sensor_fx5 = sensorlist[0]['sensor_fx5']
        # sensor_initial_data = sensorlist[0]['sensor_initial_data']

        # sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        # sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        # sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        # sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        # sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor_y = sensorlist_y[0]['sensor_gauge_factor']
        sensor_initial_data_y = sensorlist_y[0]['sensor_initial_data']
        if sensor_initial_data_y is None:
            sensor_initial_data_y = '0'
        if sensor_gauge_factor_y is None:
            sensor_gauge_factor_y = '0'

        # print(param)
        now = datetime.now()
        book = Workbook()
        # sensordata = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
        # print(sensordata)
        file_name = sensor_display_name +'-record-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

        data = [dict(r) for r in EditdataModel.sensor_data_all(param)]
        

        sensor_obj_y = SensorModel.find_by_id(sensor_idy)
        sensordetaol_obj_y = SensorDetailModel.find_by_sensor_id(sensor_idx)
        datarogger_id_y = sensor_obj_y.datarogger_id
        sensor_name_y = sensor_obj_y.sensor_name

     
        param_y= (sensor_name_y, str(datarogger_id_y), date_time_start, date_time_end)
        print("param_y",param_y)
        data_y = [dict(r) for r in EditdataModel.sensor_data_all(param_y)]
        for idx in range(len(data_y)):
            data[idx]['sensor_data_y'] = data_y[idx]['sensor_data']

        fx_len = 0
        sensor_fx_list = []
        sensor_fx_name = []

        
        # test = eval('sin(1837)')
        # print(test)
        if sensor_fx1 and len(sensor_fx1) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx1_data')
            sensor_fx_list.append('sensor_fx1_data_y')
            sensor_fx_name.append(sensor_fx1_name)
            sensor_fx_name.append(sensor_fx1_name)
            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
     
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx2_data')
            sensor_fx_list.append('sensor_fx2_data_y')
            sensor_fx_name.append(sensor_fx2_name)
            sensor_fx_name.append(sensor_fx2_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
              

                data[i]['sensor_fx2_data']=round(result1,4)
                print(data[i])
            for i in range(len(data)):

                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)

        if sensor_fx3 and len(sensor_fx3) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx3_data')
            sensor_fx_list.append('sensor_fx3_data_y')
            sensor_fx_name.append(sensor_fx3_name)
            sensor_fx_name.append(sensor_fx3_name)
            for i in range(len(data)):
               
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
             
                result1 = eval(current_fx)
                data[i]['sensor_fx3_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
           

        if sensor_fx4 and len(sensor_fx4) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx4_data')
            sensor_fx_list.append('sensor_fx4_data_y')
            sensor_fx_name.append(sensor_fx4_name)
            sensor_fx_name.append(sensor_fx4_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx4_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
                

        if sensor_fx5 and len(sensor_fx5) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx5_data')
            sensor_fx_list.append('sensor_fx5_data_y')
            sensor_fx_name.append(sensor_fx5_name)
            sensor_fx_name.append(sensor_fx5_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx5_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
     
        # rows = []       
        # for i in range(len(data)):
        #     if i == 0:
        #         rows.append("센서명","측정일시", "측정치")
        # print(pd.DataFrame([data]))
            
        if len(data) >= 0:
            sheet = book.create_sheet(sensor_name)
      
            # sheet.cell(row=1, column=1).value = "센서명"
            sheet.cell(row=1, column=1).value = "측정일시"
            sheet.cell(row=1, column=2).value = "x"
            sheet.cell(row=1, column=3).value = "y"

            for i in range(fx_len):
                sheet.cell(row=1, column=3+i+1).value = sensor_fx_name[i]
           

            for idx in range(len(data)):


                # sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_name']
                sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_data_date']
                sheet.cell(row=idx+2, column=2).value = data[idx]['sensor_data']
                sheet.cell(row=idx+2, column=3).value = data[idx]['sensor_data_y']
                for i in range(fx_len):
                    sheet.cell(row=idx+2, column=3+i+1).value = data[idx][sensor_fx_list[i]]
     
          
       
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        # print(pd.DataFrame(data))
        book.save(filename = excel_file + file_name)
       
        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

 


        return { "url" : filename }



class editdatatableallline(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('sensorgroup_type', type=str)

    def post(self):

        print("가로세로 all 테스트!!")
        params = editdatatableallline.parse.parse_args()
      

        sensorgroup_id = params["sensorgroup_id"]
        datarogger_id = params["datarogger_id"]
        date_time_start = params["date_time_start"]
        date_time_end = params["date_time_end"]
        intervalday = params["intervalday"]
        time = params["time"]
        sensorgroup_type = params["sensorgroup_type"]
  
        sensordata_list = [dict(r) for r in SensorModel.find_by_datarogger_id_snesorgroup(sensorgroup_id)]


        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_interval = sensorgroup_obj.sensorgroup_interval


        totaldatalist = []
        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
        sensorgroup_initial_date = sensorgroup_obj.sensorgroup_initial_date

        
        sensor_gauge_factor = sensorgroup_obj.sensorgroup_initial_date
        datelist = []
        for i in range(len(sensordata_list)):
            # print(sensordata_list[i]['sensor_name'])

            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']

           
            if sensor_initial_data is None or len(sensor_initial_data) == 0:
                sensor_initial_data = '0'
            if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
                sensor_gauge_factor = '0'

      

            sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(str(sensor_id))]


            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            if len(datelist) < len(data):
                datelist=[]
                for i in range(len(data)):
                    datelist.append({"sensor_data_date":data[i]['sensor_data_date']})



        for i in range(len(sensordata_list)):
            # print(sensordata_list[i]['sensor_name'])

            sensor_name = sensordata_list[i]['sensor_name']
            sensor_id = sensordata_list[i]['sensor_id']
            sensor_initial_data = sensordata_list[i]['sensor_initial_data']
            datarogger_id = sensordata_list[i]['datarogger_id']
            sensor_gauge_factor = sensordata_list[i]['sensor_gauge_factor']

           
            if sensor_initial_data is None or len(sensor_initial_data) == 0:
                sensor_initial_data = '0'
            if sensor_gauge_factor is None or len(sensor_gauge_factor) == 0:
                sensor_gauge_factor = '0'

      

            sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(str(sensor_id))]


            param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )
            data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

            # if len(datelist) < len(data):
            #     for i in range(len(data)):
            #         datelist.append({"sensor_data_date":data[i]['sensor_data_date']})

           
            if sensorlist:
                sensor_fx1 = sensorlist[0]['sensor_fx1']
                sensor_fx2 = sensorlist[0]['sensor_fx2']
                sensor_fx3 = sensorlist[0]['sensor_fx3']
                sensor_fx4 = sensorlist[0]['sensor_fx4']
                sensor_fx5 = sensorlist[0]['sensor_fx5']
                sensor_fx_check = sensorlist[0]['sensor_fx_check']


                # sensor_initial_data = sensorlist[0]['sensor_initial_data']

                sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
                sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
                sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
                sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
                sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
                # if sensor_initial_data is None:
                #     sensor_initial_data = '0'


                # param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

              
                format = "%Y-%m-%d"
                start = datetime.strptime(date_time_start[0:10], "%Y.%m.%d")
                end = datetime.strptime(date_time_end[0:10], "%Y.%m.%d")
        

                dates = []
                dates = pd.date_range(start=start, end=end, freq='1d')
                dates = dates.sort_values(ascending = False)
                sortdateall = []
                currentIndex = 0

                for i in range(len(dates)):
                    date_i = str(dates[i])
                    sortdateall.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])
                    
            
                param = (sensor_name, str(datarogger_id), date_time_start, date_time_end, '', time )

                if  len(time) ==0 and len(intervalday) !=0:
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)
                    # print(realdata)
                    # for i in range(len(realdata)):
                    #     date_i = str(realdata[i])
                    #     sortdate.append(date_i[0:4]+"."+date_i[5:7]+"."+date_i[8:10])

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1
         
                
                    sensordatas = [dict(r) for r in EditdataModel.one_day_avg(param)]
                    
                    data = []   
                    for i in range(len(sensordatas)):
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                    
                        totaldata = 0
                        totallen = 0
                        if currnetdate in sortdate:
                            # currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx

                                if currentIndex < len(sortdateall):
                                    lastdate = sortdateall[currentIndex] 
                                
                                    for j in range(len(sensordatas)):
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                        
                                            break
        

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensor_name, 'sensor_data': str(avg_data)})

              
                
                elif len(intervalday) !=0 and int(intervalday) > 1:
                    freq = intervalday+'d'
                    sortdate = []
                    realdata = pd.date_range(start=start, end=end, freq=freq)
                    realdata = realdata.sort_values(ascending = False)

                    sortdate.append(str(end)[0:4]+"."+str(end)[5:7]+"."+str(end)[8:10])
                    befor_days = datetime.now()
                    day_i = 1
                    while start < befor_days:
                        befor_days = end - timedelta(days= day_i * int(intervalday))
                        befor_one_days = str(befor_days)
                        sortdate.append(befor_one_days[0:4]+"."+befor_one_days[5:7]+"."+befor_one_days[8:10])
                        day_i += 1
     

                    sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
        

                    data = []

                    sensordatas = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                  

                    for i in range(len(sensordatas)):
              
                        currnetdate = str(sensordatas[i]['sensor_data_date'])
                 
                        totaldata = 0
                        totallen = 0
                        if currnetdate[0:10] in sortdate:
                        
                            currentdaatime = currnetdate[11:20]
                            dateindex = sortdateall.index(currnetdate[0:10])
                        
                            for idx in range(int(intervalday)):
                                currentIndex = dateindex+idx
                                if currentIndex < len(sortdateall):
                               
                                    lastdate = sortdateall[currentIndex] +" "+ currentdaatime
                                   

                                    for j in range(len(sensordatas)):
                                       
                                        if str(sensordatas[j]['sensor_data_date']) == str(lastdate):
                                            
                                            totaldata += float(sensordatas[j]['sensor_data'])
                                            totallen += 1
                                            
                                            break

                                if idx == int(intervalday)-1:
                                    avg_data = round(totaldata/totallen, 4)
                                    data.append({'sensor_data_date': sensordatas[i]['sensor_data_date'], 'sensor_name': sensordatas[i]['sensor_name'], 'sensor_data': str(avg_data)})
                    

                else:
                    # data = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]

                    data01 = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
                    datalist01 = copy.deepcopy(datelist)
                    
                    for i in range(len(datalist01)):
                        datayn = 0
                        for j in range(len(data01)):
                            
                            if datalist01[i]['sensor_data_date'] == data01[j]['sensor_data_date']:
                                datalist01[i]['sensor_name'] = data01[j]['sensor_name']
                                datalist01[i]['sensor_data'] = data01[j]['sensor_data']
                                datayn =1
                                break

                        if datayn == 0:
                            datalist01[i]['sensor_name'] = sensor_name
                            datalist01[i]['sensor_data'] = None

                    data = copy.deepcopy(datalist01)



                if sensor_fx_check == '1' and sensor_fx1 and len(sensor_fx1) != 0 :
                    
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx1
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)
                        

                            data[i]['sensor_fx_data']=str(round(result1,4))

                        else:
                            data[i]['sensor_fx_data']=None
                
                elif sensor_fx_check == '2' and sensor_fx2 and len(sensor_fx2) != 0 :
                   
                    for i in range(len(data)):

                        if data[i]['sensor_data']:
                            current_fx = sensor_fx2
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx_data']=str(round(result1,4))

                        else:
                            data[i]['sensor_fx_data']=None
            

                elif sensor_fx_check == '3' and sensor_fx3 and len(sensor_fx3) != 0 :
                  
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx3
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx_data']=str(round(result1,4))
                        else:
                            data[i]['sensor_fx_data']=None
         
                elif sensor_fx_check == '4' and sensor_fx4 and len(sensor_fx4) != 0 :
                    
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                            current_fx = sensor_fx4
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)

                            data[i]['sensor_fx_data']=str(round(result1,4))
                        else:

                            data[i]['sensor_fx_data']=None
                    
              

                elif sensor_fx_check == '5' and sensor_fx5 and len(sensor_fx5) != 0 :
                
                    for i in range(len(data)):
                        if data[i]['sensor_data']:
                    
                            current_fx = sensor_fx5
                            current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                            current_fx = current_fx.replace('$ini', sensor_initial_data)
                            current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                            result1 = eval(current_fx)
                        

                            data[i]['sensor_fx_data']=str(round(result1,4))

                        else:
                            data[i]['sensor_fx_data']=None
              

                
                totaldatalist.append(data)
                
        sensorM = []

        if(sensorgroup_type == '0202'):
            sensorM.append('0m')

        for i in range(len(sensordata_list)):
            current_inteval = int(sensorgroup_interval) * (i+1)
            sensorM.append(str(current_inteval)+'m')



        alldata = []

        if sensorgroup_type =='0201':
        # 가로형
            for idx in range(len(totaldatalist[0])):
                alldata.append({'sensor_data_date': totaldatalist[0][idx]['sensor_data_date']})

                totaldata = 0
                for j in range(len(totaldatalist)):
                    current_idx = len(totaldatalist) - 1 -j
                    if totaldatalist[current_idx][idx]['sensor_fx_data']:
                        
                        totaldata = float(totaldatalist[current_idx][idx]['sensor_fx_data'])
                       
                    alldata[idx][sensorM[current_idx]] = str(round(totaldata, 4))

        elif sensorgroup_type =='0202':
            for idx in range(len(totaldatalist[0])):
                # print(totaldatalist[0][idx]['sensor_data_date'])
                alldata.append({'sensor_data_date': totaldatalist[0][idx]['sensor_data_date']})

                totaldata = 0
                for j in range(len(sensorM)):
                    current_idx = len(sensorM) - 1 -j
                    if j == 0 :
                        alldata[idx][sensorM[current_idx]] = '0'

                    else :
                        if totaldatalist[current_idx][idx]['sensor_fx_data']:
                            totaldata += float(totaldatalist[current_idx][idx]['sensor_fx_data'])
                        alldata[idx][sensorM[current_idx]] = str(round(totaldata, 4))


             
        return {'resultCode': '0', "resultString": "SUCCESS", "data": totaldatalist, "table_fx_data": alldata}, 200



class editdatauseyn(Resource):

    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('sensor_data_date', type=str)
    parse.add_argument('datarogger_id', type=str) 

    def post(self):

        params = editdatauseyn.parse.parse_args()
        print("params", params)

        sensor_id = params["sensor_id"]
        sensor_data_date = params["sensor_data_date"]

        # sensor_obj = SensorModel.find_by_id(sensor_id)

        # datarogger_id = sensor_obj.datarogger_id

        # alarm_obj = AlarmModel.find_by_sensor_id_datetime(datarogger_id, sensor_id, str(sensor_data_date))
        # if alarm_obj is not None:
        #     print("INI:NININI")
        #     print(alarm_obj.alarm_id)
        #     alarm_obj.use_yn = 'N'
        #     alarm_obj.alarm_status = 'Y'
        #     # alarm_obj.save_to_db()


        sensor_obj = SensorModel.find_by_id(sensor_id)

        datarogger_id = str(sensor_obj.datarogger_id)

        data = [dict(r) for r in EditdataModel.find_sensor_edit_data_list(datarogger_id, sensor_data_date)]

        try:
            print("!!!!!!!")
            for i in range(len(data)):
                editdata_id = data[i]['editdata_id']
 

                editdata_obj = EditdataModel.find_by_id(editdata_id)
                sensor_name = editdata_obj.sensor_name
                print(sensor_name, datarogger_id)
                current_sensor_obj = SensorModel.find_by_datarogger_id_sensorname(sensor_name, datarogger_id)

                if current_sensor_obj is not None:
                    current_sensor_id = current_sensor_obj.sensor_id

                    print("!")
                    alarm_obj = AlarmModel.find_by_sensor_id_datetime(datarogger_id, current_sensor_id, str(sensor_data_date))
                    if alarm_obj is not None:
                        print("INI:NININI")
                        print(alarm_obj.alarm_id)
                        alarm_obj.use_yn = 'N'
                        alarm_obj.alarm_status = 'Y'
                        alarm_obj.save_to_db()

                editdata_obj.use_yn = 'N'

                editdata_obj.save_to_db()

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500


        return {'resultCode': '0', "resultString": "숨김처리 되었습니다."}, 200
        