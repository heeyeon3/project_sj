import re
import resource
from cv2 import completeSymm
from flask_jwt_extended import jwt_required
from flask_login import current_user
from flask_restful import Resource, reqparse, request
from models.company import CompanyModel
from utils.jsonutil import AlchemyEncoder as jsonEncoder
from utils.fileutil import FileUtils
from config.properties import *
import logging
import json
from models.project import ProjectModel
from models.user import UserModel
import os
import time

from datetime import datetime
from config.properties import *
from flask import send_from_directory
import requests # excel backup
from openpyxl    import load_workbook, Workbook
from openpyxl.styles    import PatternFill


class project(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('project_name', type=str)
    parse.add_argument('project_set_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('project_cost', type=str)
    parse.add_argument('project_memo', type=str)
    parse.add_argument('project_status', type=str)

    parse.add_argument('company_id', type=str)
    parse.add_argument('project_id', type=str)

    def get(self):

        project_id =              request.args.get('project_id')
        print(project_id)

        # result_string = [dict(r) for r in ProjectModel.find_by_id(project_id)]
        result_string = json.dumps(ProjectModel.find_by_id(project_id), cls=jsonEncoder)

        # result_string = [dict(r) for r in ProjectModel.find_by_company_id(project_id)]

        return {'resultCode': '0', "resultString": "Succes", "data": result_string}, 200

    def post(self):
        params = project.parse.parse_args()
        print(params)

        project_name =          params['project_name']        
        project_set_id =          params['project_set_id']        
        date_time_start =       params['date_time_start']        
        date_time_end =         params['date_time_end']        
        project_cost =          params['project_cost']        
        project_memo =          params['project_memo']        
        project_status =        params['project_status']   

        # project_total_cost = params['project_cost']  
        # project_last_date = params['date_time_end']    

        user_id = current_user.user_id    
        # current_user_id = current_user.user_id
        company_id = params['company_id']
    
        print(user_id) 

        create_date = datetime.now()
        modify_date = datetime.now()

        print(date_time_start, type(date_time_start))

        # company_obj = CompanyModel.find_by_id(company_id)

        project_set_id_obj = ProjectModel.find_by_set_id(project_set_id)



        if(project_set_id_obj):
            return {'resultCode': '0', "resultString": "프로젝트 아이디가 중복되었습니다."}, 200

        else: 

            try:
                project_obj = ProjectModel(project_name,project_set_id, date_time_start, date_time_end, project_cost, project_memo, project_status, None, None, None, None, None,None,None,
                                        'Y', user_id, company_id, create_date, modify_date )

                project_obj.save_to_db()

            except Exception as e:

                    logging.fatal(e, exc_info=True)
                    return {'resultCode': '100', "resultString": "FAIL"}, 500


            return {'resultCode': '0', "resultString": "프로젝트가 등록 되었습니다."}, 200

    def put(self):
        params = project.parse.parse_args()
        print(params)

        project_id =          params['project_id']        
        project_set_id =          params['project_set_id']        
        project_name =          params['project_name']        
        date_time_start =       params['date_time_start']        
        date_time_end =         params['date_time_end']        
        project_cost =          params['project_cost']        
        project_memo =          params['project_memo']        
        project_status =        params['project_status']   

        # project_total_cost = params['project_cost']  
        # project_last_date = params['date_time_end'] 


        user_id = current_user.user_id    
        # current_user_id = current_user.user_id
        company_id = params['company_id']
    
        print(user_id) 

        create_date = datetime.now()
        modify_date = datetime.now()

        # print(date_time_start, type(date_time_start))

        

        try:
            project_obj = ProjectModel.find_by_id(project_id)


            project_obj.project_name  = project_name
            project_obj.project_set_id  = project_set_id
            project_obj.project_st_dt  = date_time_start
            project_obj.project_ed_dt  = date_time_end
            project_obj.project_cost  = project_cost
            project_obj.project_memo  = project_memo
            project_obj.project_status  = project_status
            # project_obj.project_total_cost  = project_total_cost
            # project_obj.project_last_date  = project_last_date
            project_obj.modify_date  = modify_date
            project_obj.save_to_db()

        except Exception as e:

                logging.fatal(e, exc_info=True)
                return {'resultCode': '100', "resultString": "FAIL"}, 500


        return {'resultCode': '0', "resultString": "프로젝트 수정 되었습니다."}, 200


class projectlist(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('company_id', type=str)
    parse.add_argument('project_id', type=str)

    def get(self):

        company_id =              request.args.get('company_id')

        result_string = [dict(r) for r in ProjectModel.find_by_company_id(company_id)]

        return {'resultCode': '0', "resultString": "Succes", "data": result_string}, 200

    def post(self):

        params = projectlist.parse.parse_args()
        # print(params)

        project_id =          '0'

        print("111111")

        result_string = [dict(r) for r in ProjectModel.find_by_company_id(project_id)]
        print(result_string)
        test = [dict(r) for r in ProjectModel.project_company_list()]
        result_string.append(test[0])
        print(result_string)
        # result_string = json.dumps(ProjectModel.find_by_id(project_id), cls=jsonEncoder)

        # result_string = [dict(r) for r in ProjectModel.find_by_company_id(project_id)]

        return {'resultCode': '0', "resultString": "Succes", "data": result_string}, 200
        


class projectlistWs(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('company_id', type=str)


    def get(self):

        user_id = current_user.user_id

        user_obj = UserModel.find_by_id(user_id)
        user_gr = user_obj.user_grade

        if user_gr =='0101' or user_gr == '0102':
            company_id = user_obj.company_id 
            result_string = [dict(r) for r in ProjectModel.find_by_company_id_ws(company_id, None)]
        
        elif user_gr =='0103' or user_gr =='0104':
            project_id = user_obj.project_id 
            result_string = [dict(r) for r in ProjectModel.find_by_company_id_ws(None,project_id)]
        return {'resultCode': '0', "resultString": "Succes", "data": result_string}, 200


class projectlocation(Resource):
    parse = reqparse.RequestParser()
    parse.add_argument('company_id', type=str)

    parse.add_argument('project_lat', type=str)
    parse.add_argument('project_lng', type=str)
    parse.add_argument('project_address', type=str)
    parse.add_argument('project_img', type=str)
    parse.add_argument('weather_lat', type=str)
    parse.add_argument('weather_lng', type=str)

    parse.add_argument('project_id', type=str)


    def post(self):
        params = projectlocation.parse.parse_args()
        print(params)

        project_id =           params['project_id']        
        project_lat =          params['project_lat']        
        project_lng =          params['project_lng']        
        project_address =      params['project_address']               
        project_weather_nx =      params['weather_lat']               
        project_weather_ny =      params['weather_lng']               
      
        project_img = request.files['project_img']
        modify_date = datetime.now()

        project_obj =    ProjectModel.find_by_id(project_id)  

        try:
            project_obj = ProjectModel.find_by_id(project_id)
            company_id = project_obj.company_id

            if project_img:

                # DB 저장 필드
                # 파일명 추출 (파일명에 띄어쓰기 있는경우 파일명이 중간에 잘림)
                cont_org_nm = str(project_img.filename.replace(" ", "_"))           # 필드 생성 후 저장 해야함
                fname, ext = os.path.splitext(cont_org_nm)
                company_img_name = fname + "_"  + str(int(time.time())) + ext
                
            
                # 파일 업로드 디렉토리 세팅
                cont_path = shot_file + str(company_id) + '/' + project_id
                print(cont_path)
                try:

                    if not FileUtils.save_file(project_img, cont_path, company_img_name):
                        raise Exception('not save image %s' % cont_path + company_img_name)

                    img_url = root_url + 'shot/' + str(company_id) + '/'+project_id+ '/' + company_img_name
                    project_obj.project_img  = img_url


                except Exception as e:

                    logging.fatal(e, exc_info=True)
                    print("Screenshot FAIL")

                    return {'resultCode': '100', "resultString": "Screen Shot 파일 업로드에 실패하였습니다."}, 500

            else:
                print("파일이 존재하지 않습니다.")
                # img_url = None
                # return {'resultCode': '0', "resultString": "Screen Shot 파일이 존재하지 않습니다"}, 200

            
            # print(img_url)


            project_obj.project_lat  = project_lat
            project_obj.project_lng  = project_lng
            project_obj.project_address  = project_address
          
            project_obj.project_weather_nx  = project_weather_nx
            project_obj.project_weather_ny  = project_weather_ny
       
            
            project_obj.modify_date  = modify_date
            project_obj.save_to_db()

        except Exception as e:

                logging.fatal(e, exc_info=True)
                return {'resultCode': '100', "resultString": "FAIL"}, 500


        return {'resultCode': '0', "resultString": "프로젝트 위치 설정 되었습니다."}, 200


class currentproject(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('project_name', type=str)
    parse.add_argument('project_set_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('project_cost', type=str)
    parse.add_argument('project_memo', type=str)
    parse.add_argument('project_status', type=str)

    parse.add_argument('company_id', type=str)
    parse.add_argument('project_id', type=str)

    def get(self):

        project_id =              request.args.get('project_id')
        user_id = current_user.user_id

        user_obj = UserModel.find_by_id(user_id)
        user_gr = user_obj.user_grade
        print(user_id)
        print(user_gr)

        if user_gr == "0103" or user_gr == "0104":
            user_project_id = user_obj.project_id
            print(user_project_id)
            if str(user_project_id) != str(project_id):
                return {'resultCode': '10', "resultString": "Succes"}, 200

        elif user_gr == "0102":
            company_id = user_obj.company_id
            project_obj = ProjectModel.find_by_id(project_id)

            if str(project_obj.company_id) != str(company_id):
                return {'resultCode': '10', "resultString": "Succes"}, 200


        result_string = [dict(r) for r in ProjectModel.find_by_id_ws(project_id)]
        # result_string = json.dumps(ProjectModel.find_by_id(project_id), cls=jsonEncoder)

        # result_string = [dict(r) for r in ProjectModel.find_by_company_id(project_id)]

        return {'resultCode': '0', "resultString": "Succes", "data": result_string}, 200
    
    
    def post(self):

        project_id =              request.args.get('project_id')

        result_string = [dict(r) for r in ProjectModel.find_by_id_ws(project_id)]
        # result_string = json.dumps(ProjectModel.find_by_id(project_id), cls=jsonEncoder)

        # result_string = [dict(r) for r in ProjectModel.find_by_company_id(project_id)]

        return {'resultCode': '0', "resultString": "Succes", "data": result_string}, 200



class projectfloorplan(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('project_id', type=str)
    parse.add_argument('project_fp_img', type=str)


    def post(self):

        params = projectfloorplan.parse.parse_args()
        print(params)

        project_id =  params['project_id']
        project_fp_img = request.files['project_fp_img']

        project_obj = ProjectModel.find_by_id(project_id)
        company_id = project_obj.company_id

        

        if project_fp_img:

            # DB 저장 필드
            # 파일명 추출 (파일명에 띄어쓰기 있는경우 파일명이 중간에 잘림)
            cont_org_nm = str(project_fp_img.filename.replace(" ", "_"))           # 필드 생성 후 저장 해야함
            fname, ext = os.path.splitext(cont_org_nm)
            company_img_name = fname + "_"  + str(int(time.time())) + ext
            
        
            # 파일 업로드 디렉토리 세팅
            cont_path = shot_file + str(company_id) + '/' + project_id
            print(cont_path)
            try:

                if not FileUtils.save_file(project_fp_img, cont_path, company_img_name):
                    raise Exception('not save image %s' % cont_path + company_img_name)


                img_url = root_url + 'shot/' + str(company_id) + '/'+project_id+ '/' + company_img_name
                project_obj.project_fp_img = img_url
                print(img_url)

                project_obj.save_to_db()

            except Exception as e:

                logging.fatal(e, exc_info=True)
                print("Screenshot FAIL")

                return {'resultCode': '100', "resultString": "Screen Shot 파일 업로드에 실패하였습니다."}, 500
        else:
            print("파일이 존재하지 않습니다.")

            # return {'resultCode': '0', "resultString": "Screen Shot 파일이 존재하지 않습니다"}, 200

        

        # project_obj.project_fp_img = img_url

        # project_obj.save_to_db()

       

        return {'resultCode': '0', "resultString": "도면이미지를 등록하였습니다."}, 200
        

    


          
class projectcostexceldown(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('datarogger_id', type=str)
  

    def post(self):

        print("들어옴!!!!!!!!!!!!!!!!")

        jsonData = request.get_json()
        print(jsonData)
        print(len(jsonData))
        print(jsonData[0]['create_date'][0:4])
      


     

        # print(param)
        now = datetime.now()
        book = Workbook()
        
        # print(sensordata)
        filename = "!"
        file_name = jsonData[0]['create_date'][0:4] +'-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

            
        if len(jsonData) >= 0:
            sheet = book.create_sheet(jsonData[0]['create_date'][0:4])
      
            sheet.cell(row=1, column=1).value = "년도"
            sheet.cell(row=1, column=2).value = "등록 일시"
            sheet.cell(row=1, column=3).value = "고객사명"
            sheet.cell(row=1, column=4).value = "프로젝트명"
            sheet.cell(row=1, column=5).value = "계약 기간"
            sheet.cell(row=1, column=6).value = "구분"
            sheet.cell(row=1, column=7).value = "금액"
            sheet.cell(row=1, column=8).value = "상태"
          


            for idx in range(len(jsonData)):


                sheet.cell(row=idx+2, column=1).value = jsonData[0]['create_date'][0:4]
                sheet.cell(row=idx+2, column=2).value = jsonData[idx]['create_date']
                sheet.cell(row=idx+2, column=3).value = jsonData[idx]['company_name']
                sheet.cell(row=idx+2, column=4).value = jsonData[idx]['project_name']
                sheet.cell(row=idx+2, column=5).value = str(jsonData[idx]['start_day'][0:10]) +"~"+ str(jsonData[idx]['end_day'][0:10])
                
                if jsonData[idx]['type'] == "new":
                    sheet.cell(row=idx+2, column=6).value = "신규"
                elif jsonData[idx]['type'] == "add":
                    sheet.cell(row=idx+2, column=6).value = "연장"
                
                
                sheet.cell(row=idx+2, column=7).value = jsonData[idx]['cost']

                if jsonData[idx]['project_status'] == "N":
                    sheet.cell(row=idx+2, column=8).value = "정상"
                elif jsonData[idx]['project_status'] == "C":
                    sheet.cell(row=idx+2, column=8).value = "완료"
                elif jsonData[idx]['project_status'] == "W":
                    sheet.cell(row=idx+2, column=8).value = "대기"
          


   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        print(excel_file+file_name)
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }
