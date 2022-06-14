from flask_jwt_extended import jwt_required
from flask_restful import Resource, reqparse, request
from models.editdata import EditdataModel
from utils.jsonutil import AlchemyEncoder as jsonEncoder
import json
from models.rawdata import RawdataModel
from models.sensordetail import SensorDetailModel
from models.sensor import SensorModel

from datetime import datetime

import logging
from flask import send_from_directory
import requests # excel backup
from openpyxl    import load_workbook, Workbook
from openpyxl.styles    import PatternFill
from config.properties import *

from math import pi, sin, cos, tan, log, e
import math
import pandas as pd
# class Rawadata(Resource):

#     def post(self):
#         lastdata = RawdataModel.find_by_id_last_data_date()

#         if(lastdata):
#             return
#         else:
#             try:
#                 rawdata_obj = RawdataModel(sensor_data_date, sensor_name, sensor_data, datarogger_id, create_date, mofify_date)
                
#                 rawdata_obj.save_to_db()

#             except Exception as e:

#                 logging.fatal(e, exc_info=True)
#                 return {'resultCode': '100', "resultString": "FAIL"}, 500


#             return {'resultCode': '0', "resultString": " 등록 되었습니다."}, 200


class rawdataSensorlist(Resource):

    parse = reqparse.RequestParser()

    parse.add_argument('datarogger_id', type=str)

    def get(self):

        datarogger_id =              request.args.get('datarogger_id')

        print(datarogger_id)
        data = [dict(r) for r in RawdataModel.datarogger_sensor_list(datarogger_id)]

        return {'resultCode': '0', "resultString": "SUCCESS", "data":data}, 200

          

class rawdatagraph(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
    parse.add_argument('intervalday', type=str)
    parse.add_argument('time', type=str)
    parse.add_argument('sensor_name', type=str)

    def post(self):

        print("들어옴")
        params = rawdatagraph.parse.parse_args()
        # print(params)

        sensor_name = params["sensor_name"]
        datarogger_id = params["datarogger_id"]
        date_time_start = params["date_time_start"]
        date_time_end = params["date_time_end"]
        intervalday = params["intervalday"]
        time = params["time"]
    

        param = (sensor_name, datarogger_id, date_time_start, date_time_end, intervalday, time )

        # print(param)

        data = [dict(r) for r in RawdataModel.rawdata_sensor_name_dataroggerid(param)]

        # print(data)

      

        return {'resultCode': '0', "resultString": "SUCCESS" ,"data":data}, 200



class exceldownOrigin(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_name', type=str)

    def get(self):

        print("들어옴!!!!!!!!!!!!!!!!")

 

        sensor_id =              request.args.get('sensor_id')
        date_time_start =              request.args.get('date_time_start')
        date_time_end =              request.args.get('date_time_end')
        print(sensor_id,date_time_start,date_time_end)

        sensor_obj = SensorModel.find_by_id(sensor_id)

        sensor_name = sensor_obj.sensor_name
        datarogger_id = sensor_obj.datarogger_id
        sensor_display_name = sensor_obj.sensor_display_name
     
       

        # print(sensor_name)
        # print(datarogger_id)

        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_id)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']
        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'

     
        param = (sensor_name, str(datarogger_id),date_time_start,date_time_end)

        # print(param)
        now = datetime.now()
        book = Workbook()

        # print(sensordata)
        file_name = sensor_display_name +'-origin-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

            
        data = [dict(r) for r in RawdataModel.sensor_data_all(param)]
        # print(data)

        fx_len = 0
        sensor_fx_list = []
        sensor_fx_name = []

        if sensor_fx1 and len(sensor_fx1) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx1_data')
            sensor_fx_name.append(sensor_fx1_name)
            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data']=round(result1,4)
     
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx2_data')
            sensor_fx_name.append(sensor_fx2_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
              

                data[i]['sensor_fx2_data']=round(result1,4)
        

        if sensor_fx3 and len(sensor_fx3) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx3_data')
            sensor_fx_name.append(sensor_fx3_name)
            for i in range(len(data)):
               
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
             
                result1 = eval(current_fx)
                data[i]['sensor_fx3_data']=round(result1,4)
           

        if sensor_fx4 and len(sensor_fx4) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx4_data')
            sensor_fx_name.append(sensor_fx4_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx4_data']=round(result1,4)
                

        if sensor_fx5 and len(sensor_fx5) != 0 :
            fx_len +=1
            sensor_fx_list.append('sensor_fx5_data')
            sensor_fx_name.append(sensor_fx5_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx5_data']=round(result1,4)
      
            
        if len(data) >= 0:
            sheet = book.create_sheet(sensor_name)
      
            sheet.cell(row=1, column=1).value = "센서명"
            sheet.cell(row=1, column=2).value = "측정일시"
            sheet.cell(row=1, column=3).value = "측정치"

            for i in range(fx_len):
                sheet.cell(row=1, column=3+i+1).value = sensor_fx_name[i]
 
        

            for idx in range(len(data)):


                sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_name']
                sheet.cell(row=idx+2, column=2).value = data[idx]['sensor_data_date']
                sheet.cell(row=idx+2, column=3).value = data[idx]['sensor_data']
                for i in range(fx_len):
                    sheet.cell(row=idx+2, column=3+i+1).value = data[idx][sensor_fx_list[i]]
                    
            
       
          




        # print(data)
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        print(excel_file+file_name)
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }





class exceldownscatterorigin(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_idx', type=str)
    parse.add_argument('sensor_idy', type=str)
    parse.add_argument('date_time_start', type=str)
    parse.add_argument('date_time_end', type=str)
   



    def post(self):

        print("!!!!들어옴!!!!!!!!!!!!!!!!")
        params = exceldownscatterorigin.parse.parse_args()
     
        sensor_idx =              params['sensor_idx']
        sensor_idy =              params['sensor_idy']
        date_time_start =              params['date_time_start']
        date_time_end =              params['date_time_end']

        print(params)

        print(sensor_idx)
        print(sensor_idy)
      

        sensor_obj = SensorModel.find_by_id(sensor_idx)
        sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_idx)
        datarogger_id = sensor_obj.datarogger_id
        sensor_name = sensor_obj.sensor_name
        sensor_display_name = sensor_obj.sensor_display_name
     
        param = (sensor_name, str(datarogger_id),date_time_start ,date_time_end)
    
        sensorlist = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_idx)]
        sensor_fx1 = sensorlist[0]['sensor_fx1']
        sensor_fx2 = sensorlist[0]['sensor_fx2']
        sensor_fx3 = sensorlist[0]['sensor_fx3']
        sensor_fx4 = sensorlist[0]['sensor_fx4']
        sensor_fx5 = sensorlist[0]['sensor_fx5']
        sensor_initial_data = sensorlist[0]['sensor_initial_data']

        sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor = sensorlist[0]['sensor_gauge_factor']
        if sensor_initial_data is None:
            sensor_initial_data = '0'
        if sensor_gauge_factor is None:
            sensor_gauge_factor = '0'


        sensorlist_y = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_idy)]
        # sensor_fx1 = sensorlist[0]['sensor_fx1']
        # sensor_fx2 = sensorlist[0]['sensor_fx2']
        # sensor_fx3 = sensorlist[0]['sensor_fx3']
        # sensor_fx4 = sensorlist[0]['sensor_fx4']
        # sensor_fx5 = sensorlist[0]['sensor_fx5']
        # sensor_initial_data = sensorlist[0]['sensor_initial_data']

        # sensor_fx1_name = sensorlist[0]['sensor_fx1_name']
        # sensor_fx2_name = sensorlist[0]['sensor_fx2_name']
        # sensor_fx3_name = sensorlist[0]['sensor_fx3_name']
        # sensor_fx4_name = sensorlist[0]['sensor_fx4_name']
        # sensor_fx5_name = sensorlist[0]['sensor_fx5_name']
        sensor_gauge_factor_y = sensorlist_y[0]['sensor_gauge_factor']
        sensor_initial_data_y = sensorlist_y[0]['sensor_initial_data']
        if sensor_initial_data_y is None:
            sensor_initial_data_y = '0'
        if sensor_gauge_factor_y is None:
            sensor_gauge_factor_y = '0'

        # print(param)
        now = datetime.now()
        book = Workbook()
        # sensordata = [dict(r) for r in EditdataModel.editdata_sensor_name_dataroggerid(param)]
        # print(sensordata)
        file_name = sensor_display_name +'-origin-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

        data = [dict(r) for r in RawdataModel.sensor_data_all(param)]
        

        sensor_obj_y = SensorModel.find_by_id(sensor_idy)
        sensordetaol_obj_y = SensorDetailModel.find_by_sensor_id(sensor_idx)
        datarogger_id_y = sensor_obj_y.datarogger_id
        sensor_name_y = sensor_obj_y.sensor_name

     
        param_y= (sensor_name_y, str(datarogger_id_y),date_time_start, date_time_end )
        print("param_y",param_y)
        data_y = [dict(r) for r in RawdataModel.sensor_data_all(param_y)]
        for idx in range(len(data_y)):
            data[idx]['sensor_data_y'] = data_y[idx]['sensor_data']

        fx_len = 0
        sensor_fx_list = []
        sensor_fx_name = []

        
        # test = eval('sin(1837)')
        # print(test)
        if sensor_fx1 and len(sensor_fx1) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx1_data')
            sensor_fx_list.append('sensor_fx1_data_y')
            sensor_fx_name.append(sensor_fx1_name)
            sensor_fx_name.append(sensor_fx1_name)
            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx1
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
     
        
        if sensor_fx2 and len(sensor_fx2) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx2_data')
            sensor_fx_list.append('sensor_fx2_data_y')
            sensor_fx_name.append(sensor_fx2_name)
            sensor_fx_name.append(sensor_fx2_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
               
                result1 = eval(current_fx)
              

                data[i]['sensor_fx2_data']=round(result1,4)
                print(data[i])
            for i in range(len(data)):

                current_fx = sensor_fx2
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)

        if sensor_fx3 and len(sensor_fx3) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx3_data')
            sensor_fx_list.append('sensor_fx3_data_y')
            sensor_fx_name.append(sensor_fx3_name)
            sensor_fx_name.append(sensor_fx3_name)
            for i in range(len(data)):
               
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
             
                result1 = eval(current_fx)
                data[i]['sensor_fx3_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx3
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
           

        if sensor_fx4 and len(sensor_fx4) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx4_data')
            sensor_fx_list.append('sensor_fx4_data_y')
            sensor_fx_name.append(sensor_fx4_name)
            sensor_fx_name.append(sensor_fx4_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx4_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx4
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
                

        if sensor_fx5 and len(sensor_fx5) != 0 :
            fx_len +=2
            sensor_fx_list.append('sensor_fx5_data')
            sensor_fx_list.append('sensor_fx5_data_y')
            sensor_fx_name.append(sensor_fx5_name)
            sensor_fx_name.append(sensor_fx5_name)
            for i in range(len(data)):
                
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data'])
                current_fx = current_fx.replace('$ini', sensor_initial_data)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor)
                
                result1 = eval(current_fx)
                data[i]['sensor_fx5_data']=round(result1,4)

            for i in range(len(data)):
                current_fx = sensor_fx5
                current_fx = current_fx.replace('$sen', data[i]['sensor_data_y'])
                current_fx = current_fx.replace('$ini', sensor_initial_data_y)
                current_fx = current_fx.replace('$GF', sensor_gauge_factor_y)
       
                result1 = eval(current_fx)
   

                data[i]['sensor_fx1_data_y']=round(result1,4)
     
        # rows = []       
        # for i in range(len(data)):
        #     if i == 0:
        #         rows.append("센서명","측정일시", "측정치")
        # print(pd.DataFrame([data]))
            
        if len(data) >= 0:
            sheet = book.create_sheet(sensor_name)
      
            # sheet.cell(row=1, column=1).value = "센서명"
            sheet.cell(row=1, column=1).value = "측정일시"
            sheet.cell(row=1, column=2).value = "x"
            sheet.cell(row=1, column=3).value = "y"

            for i in range(fx_len):
                sheet.cell(row=1, column=3+i+1).value = sensor_fx_name[i]
           

            for idx in range(len(data)):


                # sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_name']
                sheet.cell(row=idx+2, column=1).value = data[idx]['sensor_data_date']
                sheet.cell(row=idx+2, column=2).value = data[idx]['sensor_data']
                sheet.cell(row=idx+2, column=3).value = data[idx]['sensor_data_y']
                for i in range(fx_len):
                    sheet.cell(row=idx+2, column=3+i+1).value = data[idx][sensor_fx_list[i]]
     
          
       
   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        # print(pd.DataFrame(data))
        book.save(filename = excel_file + file_name)
       
        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        print(excel_file+file_name)
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

 


        return { "url" : filename }