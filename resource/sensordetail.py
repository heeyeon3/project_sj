from argparse import Action
from ast import Param
from datetime import datetime
from fileinput import filename
from flask_jwt_extended import jwt_required
from flask_login import current_user
from flask_restful import Resource, reqparse, request
from sqlalchemy import null
from models.sensorgroup import SensorgroupModel
from utils.jsonutil import AlchemyEncoder as jsonEncoder
import json
from models.rawdata import RawdataModel
from models.editdata import EditdataModel
from models.sensor import SensorModel
from models.sensordetail import SensorDetailModel
import logging
from config.properties import *
from flask import send_from_directory
import requests # excel backup
from openpyxl    import load_workbook, Workbook
from openpyxl.styles    import PatternFill

from math import pi, sin, cos, tan, log, e
import math
import random



class Sensordetail(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_gl1_max', type=str)
    parse.add_argument('sensor_gl1_min', type=str)
    parse.add_argument('sensor_gl2_max', type=str)
    parse.add_argument('sensor_gl2_min', type=str)
    parse.add_argument('sensor_gl3_max', type=str)
    parse.add_argument('sensor_gl3_min', type=str)
    parse.add_argument('sensor_fx1', type=str)
    parse.add_argument('sensor_fx2', type=str)
    parse.add_argument('sensor_fx3', type=str)
    parse.add_argument('sensor_fx4', type=str)
    parse.add_argument('sensor_fx5', type=str)
    parse.add_argument('sensor_max', type=str)
    parse.add_argument('sensor_min', type=str)
    
    parse.add_argument('sensor_fx1_name', type=str)
    parse.add_argument('sensor_fx2_name', type=str)
    parse.add_argument('sensor_fx3_name', type=str)
    parse.add_argument('sensor_fx4_name', type=str)
    parse.add_argument('sensor_fx5_name', type=str)
    
    parse.add_argument('sensor_fx1_id', type=str)
    parse.add_argument('sensor_fx2_id', type=str)
    parse.add_argument('sensor_fx3_id', type=str)
    parse.add_argument('sensor_fx4_id', type=str)
    parse.add_argument('sensor_fx5_id', type=str)

    parse.add_argument('sensor_weight', type=str)
    parse.add_argument('sensor_deviation', type=str)
    parse.add_argument('sensor_st_over_ex', type=str)
    parse.add_argument('sensor_st_over_wt', type=str)
    parse.add_argument('sensor_dev_over_ex', type=str)
    parse.add_argument('sensor_dev_over_wt', type=str)
    parse.add_argument('sensor_null_ex', type=str)
    parse.add_argument('sensor_default_wt', type=str)
    parse.add_argument('sensor_noti', type=str)

    parse.add_argument('sensor_detail_id', type=str)
    parse.add_argument('sensor_id', type=str)
    parse.add_argument('sensorgroup_id', type=str)

    def get(self):

        sensor_id = request.args.get('sensor_id')
        datarogger_id = request.args.get('datarogger_id')


        # user_id = current_user.user_id
        data = [dict(r) for r in SensorDetailModel.sensordetail_info_sensor_id(sensor_id)]
        print("data", data)

        # print(root_url)
        # print(root_url.split('/'))
        rootsplit = root_url.split('/')
        rootsplit = "/".join(rootsplit[0:3])
        print("/".join(rootsplit[0:3]))
        if  data[0]['sensorgroup_type'] and (data[0]['sensorgroup_type'] == '0201' or data[0]['sensorgroup_type'] == '0202' or data[0]['sensorgroup_type'] == '0204'or data[0]['sensorgroup_type'] == '0205' or data[0]['sensorgroup_type'] == '0206'):
            sensor_url = str(rootsplit) +"/ws-02-2-1?c2V="+str(data[0]['sensor_id'])+"&cHJ="+str(data[0]['project_id'])+"&Hlw="+str(data[0]['sensorgroup_type'])
            # sensor_url = str(rootsplit) +"/ws-02-2-1?sensor_id="+str(data[0]['sensor_id'])+"&datarogger_id="+str(data[0]['datarogger_id'])+"&sensor_name="+str(data[0]['sensor_name'])+"&project_id="+str(data[0]['project_id'])+"&sensorgroup_type="+str(data[0]['sensorgroup_type'])
            data[0]['sensor_url'] = sensor_url
        elif data[0]['sensorgroup_type'] and data[0]['sensorgroup_type'] == '0203':
            #sensor_idx=6&sensor_idy=1&sensor_namex=2P02&sensor_namey=1P01&project_id=1&sensorgroup_type=0203

            sensorgroup_id = data[0]['sensorgroup_id']
            sensor_display_name = data[0]['sensor_display_name']
            param = (sensor_display_name, str(sensorgroup_id), sensor_id)
            anorther_sensor = [dict(r) for r in SensorDetailModel.find_by_scatter_another_sensor(param)]

            another_sensor_id = str(anorther_sensor[0]['sensor_id'])
  
            another_data = [dict(r) for r in SensorDetailModel.sensordetail_info_sensor_id(another_sensor_id)]
        
            if anorther_sensor and len(anorther_sensor) > 0 :
                another_sensor_id = str(anorther_sensor[0]['sensor_id'])
                another_data = [dict(r) for r in SensorDetailModel.sensordetail_info_sensor_id(another_sensor_id)]
                if another_data and data[0]['sensor_type'] == 'x':
                    
                    sensor_url = str(rootsplit) +"/ws-02-2-6?c2Vx="+str(data[0]['sensor_id'])+"&c2Vy="+str(another_data[0]['sensor_id'])+"&cHJ="+str(data[0]['project_id'])+"&Hlw=0203"
                elif another_data and data[0]['sensor_type'] == 'y':
                    
                    sensor_url = str(rootsplit) +"/ws-02-2-6?c2Vx="+str(another_data[0]['sensor_id'])+"&c2Vy="+str(data[0]['sensor_id'])+"&cHJ="+str(data[0]['project_id'])+"&Hlw=0203"
                
                else :
                    data[0]['sensor_url'] = None

                data[0]['sensor_url'] = sensor_url
            else:
                data[0]['sensor_url'] = None
        else:
            data[0]['sensor_url'] = None
       

        return {'resultCode': '0', "resultString": "SUCCESS", "data":data}, 200
       



    def post(self):
        print("===============================test===============================")
        params = Sensordetail.parse.parse_args()
        print(params)
        
        # sensor_id = request.args.get('sensor_id')
        # datarogger_id = request.args.get('datarogger_id')

    

        sensor_detail_id = params['sensor_detail_id']
        sensor_gl1_max = params['sensor_gl1_max']
        sensor_gl1_min = params['sensor_gl1_min']
        sensor_gl2_max = params['sensor_gl2_max']
        sensor_gl2_min = params['sensor_gl2_min']
        sensor_gl3_max = params['sensor_gl3_max']
        sensor_gl3_min = params['sensor_gl3_min']
        sensor_fx1 = params['sensor_fx1']
        sensor_fx2 = params['sensor_fx2']
        sensor_fx3 = params['sensor_fx3']
        sensor_fx4 = params['sensor_fx4']
        sensor_fx5 = params['sensor_fx5']
        sensor_max = params['sensor_max']
        sensor_min = params['sensor_min']

        sensor_fx1_name = params['sensor_fx1_name']
        sensor_fx2_name = params['sensor_fx2_name']
        sensor_fx3_name = params['sensor_fx3_name']
        sensor_fx4_name = params['sensor_fx4_name']
        sensor_fx5_name = params['sensor_fx5_name']
        
        sensor_fx1_id = params['sensor_fx1_id']
        sensor_fx2_id = params['sensor_fx2_id']
        sensor_fx3_id = params['sensor_fx3_id']
        sensor_fx4_id = params['sensor_fx4_id']
        sensor_fx5_id = params['sensor_fx5_id']
        
        sensor_weight = params['sensor_weight']
        sensor_deviation = params['sensor_deviation']

        sensor_st_over_ex = params['sensor_st_over_ex']
        sensor_st_over_wt = params['sensor_st_over_wt']
        sensor_dev_over_ex = params['sensor_dev_over_ex']
        sensor_dev_over_wt = params['sensor_dev_over_wt']
        sensor_null_ex = params['sensor_null_ex']
        sensor_default_wt = params['sensor_default_wt']
        
        sensor_noti = params['sensor_noti']
        sensor_id = params['sensor_id']
        sensorgroup_id = params['sensorgroup_id']
        datarogger_id = params['datarogger_id']

        use_yn = 'Y'
        user_id = current_user.user_id
        create_date = datetime.now()
        modify_date = datetime.now()
     
        print("sensor_noti", sensor_noti)

        if sensor_fx1_name == "선택":
            sensor_fx1_name = None
        if sensor_fx2_name == "선택":
            sensor_fx2_name = None
        if sensor_fx3_name == "선택":
            sensor_fx3_name = None
        if sensor_fx4_name == "선택":
            sensor_fx4_name = None
        if sensor_fx5_name == "선택":
            sensor_fx5_name = None


        if sensor_fx1 == "undefined":
            sensor_fx1 = None
        if sensor_fx2 == "undefined":
            sensor_fx2 = None
        if sensor_fx3 == "undefined":
            sensor_fx3 = None
        if sensor_fx4 == "undefined":
            sensor_fx4 = None
        if sensor_fx5 == "undefined":
            sensor_fx5 = None

        # sensor_name = SensorModel.find_by_id(sensor_id).sensor_name
        # data = [dict(r) for r in RawdataModel.sensor_data_first(datarogger_id, sensor_name)]
        # print(data)

        # sensor_initial_data = data[0]['sensor_data']
        # sensor_initial_date = data[0]['sensor_data_date']
        # print(sensor_initial_data, sensor_initial_date)

        



        if len(sensor_gl1_max) ==0: sensor_gl1_max = None
        if len(sensor_gl1_min) ==0: sensor_gl1_min = None
        if len(sensor_gl2_max) ==0: sensor_gl2_max = None
        if len(sensor_gl2_min) ==0: sensor_gl2_min = None
        if len(sensor_gl3_max) ==0: sensor_gl3_max = None
        if len(sensor_gl3_min) ==0: sensor_gl3_min = None
        # if len(sensor_fx1) ==0: sensor_fx1 = None
        # if len(sensor_fx2) ==0: sensor_fx2 = None
        # if len(sensor_fx3) ==0: sensor_fx3 = None
        # if len(sensor_fx4) ==0: sensor_fx4 = None
        # if len(sensor_fx5) ==0: sensor_fx5 = None
        if len(sensor_max) ==0: sensor_max = None
        if len(sensor_min) ==0: sensor_min = None
        if len(sensor_weight) ==0: sensor_weight = None
        if len(sensor_deviation) ==0: sensor_deviation = None
        if len(sensor_st_over_ex) ==0: sensor_st_over_ex = None
        if len(sensor_st_over_wt) ==0: sensor_st_over_wt = None
        if len(sensor_dev_over_ex) ==0: sensor_dev_over_ex = None
        if len(sensor_dev_over_wt) ==0: sensor_dev_over_wt = None
        if len(sensor_null_ex) ==0: sensor_null_ex = None
        if len(sensor_default_wt) ==0: sensor_default_wt = None
        if len(sensor_noti) ==0: sensor_noti = None
        if sensor_noti == 'null': 
            print("!!@#!@#")
            sensor_noti = None

        if len(sensor_fx1) !=0: sensor_fx_check = '1'
        elif len(sensor_fx2) != 0: sensor_fx_check = '2'
        elif len(sensor_fx3) != 0: sensor_fx_check = '3'
        elif len(sensor_fx4) != 0: sensor_fx_check = '4'
        elif len(sensor_fx5) != 0: sensor_fx_check = '5'
        else : sensor_fx_check = None

        
        try:
           
            if(len(sensor_detail_id)==0):
                print("not sensordetail id")
                sensor_name = SensorModel.find_by_id(sensor_id).sensor_name
                data = [dict(r) for r in RawdataModel.sensor_data_first(datarogger_id, sensor_name)]
                print(data)

                if(len(data) != 0):
                    sensor_initial_data = data[0]['sensor_data']
                    sensor_initial_date = data[0]['sensor_data_date']

                else: 
                    print("not data")
                    sensor_initial_data = None
                    sensor_initial_date = None

                sensordetail_obj = SensorDetailModel(sensor_initial_date,sensor_initial_data, sensor_gl1_max, sensor_gl1_min, sensor_gl2_max, sensor_gl2_min, sensor_gl3_max, sensor_gl3_min, sensor_fx1, sensor_fx2, sensor_fx3, sensor_fx4, sensor_fx5,
                                sensor_max, sensor_min, sensor_weight, sensor_deviation, sensor_st_over_ex, sensor_st_over_wt, sensor_dev_over_ex, sensor_dev_over_wt, sensor_null_ex, sensor_default_wt, sensor_noti, '0.002', sensor_fx1_name,  sensor_fx2_name,  sensor_fx3_name,  sensor_fx4_name,  sensor_fx5_name,  sensor_fx_check,  
                                sensor_fx1_id, sensor_fx2_id, sensor_fx3_id, sensor_fx4_id, sensor_fx5_id, use_yn, user_id,
                                sensor_id, create_date, modify_date)



                if(len(sensorgroup_id)!=0):

                    sensor_obj = SensorModel.find_by_id(sensor_id)
                    print(sensor_obj.sensorgroup_id)
                    sensor_obj.sensorgroup_id = sensorgroup_id
                    sensor_obj.save_to_db()

                    sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
                    sensorgroup_obj.sensorgroup_initial_date = sensor_initial_date
                    sensorgroup_obj.save_to_db()
          
                sensordetail_obj.save_to_db()
                

            else:
                print("sensordetail id")
                sensordetail_obj = SensorDetailModel.find_by_id(sensor_detail_id)

                sensordetail_obj.sensor_gl1_max = sensor_gl1_max
                sensordetail_obj.sensor_gl1_min = sensor_gl1_min
                sensordetail_obj.sensor_gl2_max = sensor_gl2_max
                sensordetail_obj.sensor_gl2_min = sensor_gl2_min
                sensordetail_obj.sensor_gl3_max = sensor_gl3_max
                sensordetail_obj.sensor_gl3_min = sensor_gl3_min

                if sensordetail_obj.sensor_fx1_id != sensor_fx1_id :
                    sensordetail_obj.sensor_fx1 = sensor_fx1
                    sensordetail_obj.sensor_fx1_name = sensor_fx1_name
                    sensordetail_obj.sensor_fx1_id = sensor_fx1_id
                
                if sensordetail_obj.sensor_fx2_id != sensor_fx2_id :
                    sensordetail_obj.sensor_fx2 = sensor_fx2
                    sensordetail_obj.sensor_fx2_name = sensor_fx2_name
                    sensordetail_obj.sensor_fx2_id = sensor_fx2_id

                if sensordetail_obj.sensor_fx3_id != sensor_fx3_id :
                    sensordetail_obj.sensor_fx3 = sensor_fx3
                    sensordetail_obj.sensor_fx3_name = sensor_fx3_name
                    sensordetail_obj.sensor_fx3_id = sensor_fx3_id

                if sensordetail_obj.sensor_fx4_id != sensor_fx4_id :
                    sensordetail_obj.sensor_fx4 = sensor_fx4
                    sensordetail_obj.sensor_fx4_name = sensor_fx4_name
                    sensordetail_obj.sensor_fx4_id = sensor_fx4_id

                if sensordetail_obj.sensor_fx5_id != sensor_fx5_id :
                    sensordetail_obj.sensor_fx5 = sensor_fx5
                    sensordetail_obj.sensor_fx5_name = sensor_fx5_name
                    sensordetail_obj.sensor_fx5_id = sensor_fx5_id

                sensordetail_obj.sensor_max = sensor_max
                sensordetail_obj.sensor_min = sensor_min
                sensordetail_obj.sensor_weight = sensor_weight
                sensordetail_obj.sensor_deviation = sensor_deviation

                sensordetail_obj.sensor_st_over_ex = sensor_st_over_ex
                sensordetail_obj.sensor_st_over_wt = sensor_st_over_wt
                sensordetail_obj.sensor_dev_over_ex = sensor_dev_over_ex
                sensordetail_obj.sensor_dev_over_wt = sensor_dev_over_wt
                sensordetail_obj.sensor_null_ex = sensor_null_ex
                sensordetail_obj.sensor_default_wt = sensor_default_wt
                sensordetail_obj.sensor_noti = sensor_noti

                sensordetail_obj.save_to_db()

                if(len(sensorgroup_id)!=0):

                    sensor_obj = SensorModel.find_by_id(sensor_id)
                    print(sensor_obj.sensorgroup_id)
                    sensor_obj.sensorgroup_id = sensorgroup_id
                    sensor_obj.save_to_db()

                    sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_id)

                    sensor_initial_date = sensordetail_obj.sensor_initial_date

                    sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
                    sensorgroup_obj.sensorgroup_initial_date = sensor_initial_date
                    sensorgroup_obj.save_to_db()
          
                
                

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

    
        return {'resultCode': '0', "resultString": "SUCCESS"}, 200



class SensordetailList(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('project_id', type=str)


    def get(self):

        project_id = request.args.get('project_id')
        data = [dict(r) for r in SensorDetailModel.sensordetail_info_project_id(project_id)]

        print(root_url)

        return {'resultCode': '0', "resultString": "SUCCESS", "data": data}, 200



    def post(self):

        params = SensordetailList.parse.parse_args()
        print(params)
        # jsonData = request.get_json()

        # print(jsonData)
        datarogger_id = params['datarogger_id']

        # user_id = current_user.user_id
        data = [dict(r) for r in SensorDetailModel.sensordetail_list(datarogger_id)]
        # print(data)

        return {'resultCode': '0', "resultString": "SUCCESS", "data":data}, 200


class Sensordetailmodal(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_gl1_max', type=str)
    parse.add_argument('sensor_gl1_min', type=str)
    parse.add_argument('sensor_gl2_max', type=str)
    parse.add_argument('sensor_gl2_min', type=str)
    parse.add_argument('sensor_gl3_max', type=str)
    parse.add_argument('sensor_gl3_min', type=str)
    parse.add_argument('sensor_id', type=str)
    parse.add_argument('sensor_detail_id', type=str)


    def get(self):

        sensor_detail_id = request.args.get('sensor_detail_id')
        sensor_gauge_factor = request.args.get('sensor_gauge_factor')

        print(sensor_detail_id,sensor_gauge_factor)


        sensor_detail_obj = SensorDetailModel.find_by_id(sensor_detail_id)
        sensor_detail_obj.sensor_gauge_factor =sensor_gauge_factor
        sensor_detail_obj.save_to_db()

        # user_id = current_user.user_id
        # data = [dict(r) for r in SensorDetailModel.sensordetail_info_sensor_id(sensor_id, datarogger_id)]
        # print(data)

        return {'resultCode': '0', "resultString": "SUCCESS"}, 200

    def post(self):

        print("들어옴")
        params = Sensordetailmodal.parse.parse_args()
        print(params)

        sensor_id = params['sensor_id']

        sensor_gl1_max = params['sensor_gl1_max']
        sensor_gl1_min = params['sensor_gl1_min']
        sensor_gl2_max = params['sensor_gl2_max']
        sensor_gl2_min = params['sensor_gl2_min']
        sensor_gl3_max = params['sensor_gl3_max']
        sensor_gl3_min = params['sensor_gl3_min']

       
        if sensor_gl1_max is not None and len(sensor_gl1_max) ==0: sensor_gl1_max = None
        else : sensor_gl1_min = "-"+sensor_gl1_max
        # if sensor_gl1_min and len(sensor_gl1_min) ==0: sensor_gl1_min = None
        if sensor_gl2_max is not None and len(sensor_gl2_max) ==0: sensor_gl2_max = None
        else : sensor_gl2_min = "-"+sensor_gl2_max
        # if sensor_gl2_min and len(sensor_gl2_min) ==0: sensor_gl2_min = None
        if sensor_gl3_max is not None and len(sensor_gl3_max) ==0: 
            sensor_gl3_max = None
            print("!!")
        else : sensor_gl3_min = "-"+sensor_gl3_max

      

        # print(jsonData)
        sensor_detail_id = SensorDetailModel.find_by_sensor_id(sensor_id).sensor_detail_id
    
        sensor_detail_obj = SensorDetailModel.find_by_id(sensor_detail_id)

        sensor_detail_obj.sensor_gl1_max = sensor_gl1_max
        sensor_detail_obj.sensor_gl1_min = sensor_gl1_min
        sensor_detail_obj.sensor_gl2_max = sensor_gl2_max
        sensor_detail_obj.sensor_gl2_min = sensor_gl2_min
        sensor_detail_obj.sensor_gl3_max = sensor_gl3_max
        sensor_detail_obj.sensor_gl3_min = sensor_gl3_min


       
        sensor_detail_obj.save_to_db()

        return {'resultCode': '0', "resultString": "SUCCESS"}, 200


class Sensordetailmodalscatter(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_gl1_max', type=str)
    parse.add_argument('sensor_gl1_min', type=str)
    parse.add_argument('sensor_gl2_max', type=str)
    parse.add_argument('sensor_gl2_min', type=str)
    parse.add_argument('sensor_gl3_max', type=str)
    parse.add_argument('sensor_gl3_min', type=str)
    parse.add_argument('sensor_idx', type=str)
    parse.add_argument('sensor_idy', type=str)
    parse.add_argument('sensor_detail_id', type=str)


    def get(self):

        sensor_detail_id = request.args.get('sensor_detail_id')
        sensor_gauge_factor = request.args.get('sensor_gauge_factor')

        print(sensor_detail_id,sensor_gauge_factor)


        sensor_detail_obj = SensorDetailModel.find_by_id(sensor_detail_id)
        sensor_detail_obj.sensor_gauge_factor =sensor_gauge_factor
        # sensor_detail_obj.save_to_db()

        # user_id = current_user.user_id
        # data = [dict(r) for r in SensorDetailModel.sensordetail_info_sensor_id(sensor_id, datarogger_id)]
        # print(data)

        return {'resultCode': '0', "resultString": "SUCCESS"}, 200

    def post(self):

        print("들어옴")
        params = Sensordetailmodalscatter.parse.parse_args()
        print(params)

        sensor_idx = params['sensor_idx']
        sensor_idy = params['sensor_idy']

        sensor_gl1_max = params['sensor_gl1_max']
        sensor_gl1_min = params['sensor_gl1_min']
        sensor_gl2_max = params['sensor_gl2_max']
        sensor_gl2_min = params['sensor_gl2_min']
        sensor_gl3_max = params['sensor_gl3_max']
        sensor_gl3_min = params['sensor_gl3_min']

    

        if sensor_gl1_max is not None and len(sensor_gl1_max) ==0: sensor_gl1_max = None
        else : sensor_gl1_min = "-"+sensor_gl1_max
        # if sensor_gl1_min and len(sensor_gl1_min) ==0: sensor_gl1_min = None
        if sensor_gl2_max is not None and len(sensor_gl2_max) ==0: sensor_gl2_max = None
        else : sensor_gl2_min = "-"+sensor_gl2_max
        # if sensor_gl2_min and len(sensor_gl2_min) ==0: sensor_gl2_min = None
        if sensor_gl3_max is not None and len(sensor_gl3_max) ==0: 
            sensor_gl3_max = None
            print("!!")
        else : sensor_gl3_min = "-"+sensor_gl3_max
      

        # print(jsonData)
        sensor_detail_idx = SensorDetailModel.find_by_sensor_id(sensor_idx).sensor_detail_id
    
        sensor_detail_objx = SensorDetailModel.find_by_id(sensor_detail_idx)

        sensor_detail_objx.sensor_gl1_max = sensor_gl1_max
        sensor_detail_objx.sensor_gl1_min = sensor_gl1_min
        sensor_detail_objx.sensor_gl2_max = sensor_gl2_max
        sensor_detail_objx.sensor_gl2_min = sensor_gl2_min
        sensor_detail_objx.sensor_gl3_max = sensor_gl3_max
        sensor_detail_objx.sensor_gl3_min = sensor_gl3_min

        sensor_detail_idy = SensorDetailModel.find_by_sensor_id(sensor_idy).sensor_detail_id
    
        sensor_detail_objy = SensorDetailModel.find_by_id(sensor_detail_idy)

        sensor_detail_objy.sensor_gl1_max = sensor_gl1_max
        sensor_detail_objy.sensor_gl1_min = sensor_gl1_min
        sensor_detail_objy.sensor_gl2_max = sensor_gl2_max
        sensor_detail_objy.sensor_gl2_min = sensor_gl2_min
        sensor_detail_objy.sensor_gl3_max = sensor_gl3_max
        sensor_detail_objy.sensor_gl3_min = sensor_gl3_min


        try:
            sensor_detail_objx.save_to_db()
            sensor_detail_objy.save_to_db()

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

    
        return {'resultCode': '0', "resultString": "SUCCESS"}, 200

      





# class Sensorinitialmodal(Resource):
#     parse = reqparse.RequestParser()
 

#     parse.add_argument('sensor_id', type=str)
#     parse.add_argument('datarogger_id', type=str)
#     parse.add_argument('sensor_name', type=str)
#     parse.add_argument('sensor_initial_date', type=str)
#     parse.add_argument('sensor_name', type=str)


#     def post(self):

#         print("들어옴")
#         params = Sensorinitialmodal.parse.parse_args()
#         print(params)

#         sensor_id = params['sensor_id']

#         datarogger_id = params['datarogger_id']
#         sensor_name = params['sensor_name']
#         sensor_initial_date = params['sensor_initial_date']
       
   
#         sensor_obj = SensorModel.find_by_id(sensor_id)
#         sensorgroup_id = sensor_obj.sensorgroup_id
#         sensorlist = [dict(r) for r in SensorModel.all_sensor_list(str(sensorgroup_id))]
#         print(sensorlist)



#         try:
#             for i in range(len(sensorlist)):
#                 current_sensor_id = sensorlist[i]['sensor_id']
#                 sensor_name = sensorlist[i]['sensor_name']
#                 datarogger_id = sensorlist[i]['datarogger_id']

#                 sensor_detail_obj = SensorDetailModel.find_by_sensor_id(str(current_sensor_id))
#                 sensor_detail_obj.sensor_initial_date  = sensor_initial_date

#                 param = (sensor_name, str(datarogger_id), sensor_initial_date)
#                 sensordata = [dict(r) for r in EditdataModel.find_initial_data(param)]

#                 if(len(sensordata)>0):
#                     sensor_initial_data = sensordata[0]['sensor_data']
#                     print(sensor_initial_data)
#                     sensor_detail_obj.sensor_initial_data = sensor_initial_data
#                 else:
#                     sensor_detail_obj.sensor_initial_data = None

#                 sensor_detail_obj.save_to_db()

#             sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
#             sensorgroup_obj.sensorgroup_initial_date = sensor_initial_date
#             sensorgroup_obj.save_to_db()
#         except Exception as e:

#             logging.fatal(e, exc_info=True)
#             return {'resultCode': '100', "resultString": "FAIL"}, 500
  

#         return {'resultCode': '0', "resultString": "SUCCESS"}, 200


class Sensorinitialmodal(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_name', type=str)
    parse.add_argument('sensor_initial_date', type=str)
    parse.add_argument('sensor_name', type=str)


    def post(self):

        print("들어옴")
        params = Sensorinitialmodal.parse.parse_args()
        print(params)

        sensor_id = params['sensor_id']

        datarogger_id = params['datarogger_id']
        sensor_name = params['sensor_name']
        sensorgroup_initial_date = params['sensor_initial_date']

        
       
   
        sensor_obj = SensorModel.find_by_id(sensor_id)
        sensorgroup_id = str(sensor_obj.sensorgroup_id)
        sensorlist = [dict(r) for r in SensorModel.all_sensor_list(sensorgroup_id)]
        print(sensorlist)

        sensorgroup_initial_date_1 = sensorgroup_initial_date[0:4]+"-"+sensorgroup_initial_date[5:7]+"-"+sensorgroup_initial_date[8:10]+" "+sensorgroup_initial_date[11:16]
        print("sensorgroup_initial_date_1",sensorgroup_initial_date_1)

        sensor_list = [dict(r) for r in SensorModel.all_sensor_list(sensorgroup_id)]

        datarogger_id = str(sensor_list[0]['datarogger_id'])
        print(datarogger_id)

        sensor_first = [dict(r) for r in EditdataModel.find_sensor_first_data(datarogger_id,sensorgroup_initial_date_1)]
        print(sensor_first)
        if len(sensor_first) == 0:
            return {'resultCode': '0', "resultString": "선택하신 날짜에 데이터 존재하지 않아 설정을 완료할수 없습니다."}, 200

        else:
            sensor_initial_date =sensor_first[0]['sensor_data_date']



        try:
            for i in range(len(sensorlist)):
                current_sensor_id = sensorlist[i]['sensor_id']
                sensor_name = sensorlist[i]['sensor_name']
                datarogger_id = sensorlist[i]['datarogger_id']

                sensor_detail_obj = SensorDetailModel.find_by_sensor_id(str(current_sensor_id))
                sensor_detail_obj.sensor_initial_date  = sensor_initial_date

                param = (sensor_name, str(datarogger_id), sensor_initial_date)
                sensordata = [dict(r) for r in EditdataModel.find_initial_data(param)]

                if(len(sensordata)>0):
                    sensor_initial_data = sensordata[0]['sensor_data']
                    print(sensor_initial_data)
                    sensor_detail_obj.sensor_initial_data = sensor_initial_data
                else:
                    sensor_detail_obj.sensor_initial_data = None

                sensor_detail_obj.save_to_db()

            sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
            sensorgroup_obj.sensorgroup_initial_date = sensor_initial_date
            sensorgroup_obj.save_to_db()
        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500
  

        return {'resultCode': '0', "resultString": "Initial date가 등록 되었습니다."}, 200



class Sensorinitialmodalscatter(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_idx', type=str)
    parse.add_argument('sensor_idy', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_namex', type=str)
    parse.add_argument('sensor_namey', type=str)
    parse.add_argument('sensor_initial_date', type=str)
    parse.add_argument('sensor_name', type=str)


    def post(self):

        print("들어옴")
        params = Sensorinitialmodalscatter.parse.parse_args()
        print(params)

        sensor_idx = params['sensor_idx']
        sensor_idy = params['sensor_idy']

        datarogger_id = params['datarogger_id']
        sensor_namex = params['sensor_namex']
        sensor_namey = params['sensor_namey']
        sensor_initial_date = params['sensor_initial_date']
       
   
      

        # print(jsonData)
        sensor_detail_idx = SensorDetailModel.find_by_sensor_id(sensor_idx).sensor_detail_id
        sensor_detail_idy = SensorDetailModel.find_by_sensor_id(sensor_idy).sensor_detail_id
    
        sensor_detail_objx = SensorDetailModel.find_by_id(sensor_detail_idx)
        sensor_detail_objy = SensorDetailModel.find_by_id(sensor_detail_idy)

        sensor_detail_objx.sensor_initial_date = sensor_initial_date
        sensor_detail_objy.sensor_initial_date = sensor_initial_date

        try:
            paramx = (sensor_namex, datarogger_id, sensor_initial_date)
            sensordatax = [dict(r) for r in EditdataModel.find_initial_data(paramx)]

            if(len(sensordatax)>0):
                sensor_initial_data = sensordatax[0]['sensor_data']
                print(sensor_initial_data)
                sensor_detail_objx.sensor_initial_data = sensor_initial_data
            else:
                sensor_detail_objx.sensor_initial_data = None



            paramy = (sensor_namey, datarogger_id, sensor_initial_date)
            sensordatay = [dict(r) for r in EditdataModel.find_initial_data(paramy)]

            if(len(sensordatay)>0):
                sensor_initial_data = sensordatay[0]['sensor_data']
                print(sensor_initial_data)
                sensor_detail_objy.sensor_initial_data = sensor_initial_data
            else:
                sensor_detail_objy.sensor_initial_data = None

        
            sensor_detail_objx.save_to_db()
            sensor_detail_objy.save_to_db()

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

    
        return {'resultCode': '0', "resultString": "SUCCESS"}, 200




          
class sensorlistexceldown(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('datarogger_id', type=str)
  

    def get(self):

        print("들어옴!!!!!!!!!!!!!!!!")

       
        datarogger_id =              request.args.get('datarogger_id')

 
        print(datarogger_id)


     
        rootsplit = root_url.split('/')
        rootsplit = "/".join(rootsplit[0:3])

        # print(param)
        now = datetime.now()
        book = Workbook()
        sensordata = [dict(r) for r in SensorDetailModel.sensordetail_list(datarogger_id)]
        print("sensordata", sensordata)
    
        file_name = sensordata[0]['datarogger_name'] +'-'+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"

            
        if len(sensordata) >= 0:
            sheet = book.create_sheet(sensordata[0]['datarogger_name'])
      
            sheet.cell(row=1, column=1).value = "Data no."
            sheet.cell(row=1, column=2).value = "센서명"
            sheet.cell(row=1, column=3).value = "구분"
            sheet.cell(row=1, column=4).value = "센서 S/N"
            sheet.cell(row=1, column=5).value = "설치 지점"
            sheet.cell(row=1, column=6).value = "센서 그룹"
            sheet.cell(row=1, column=7).value = "센서 url"
          


            for idx in range(len(sensordata)):


                sheet.cell(row=idx+2, column=1).value = sensordata[idx]['sensor_index']
                sheet.cell(row=idx+2, column=2).value = sensordata[idx]['sensor_name']
                sheet.cell(row=idx+2, column=3).value = sensordata[idx]['sensor_type']
                sheet.cell(row=idx+2, column=4).value = sensordata[idx]['sensor_sn']
                sheet.cell(row=idx+2, column=5).value = sensordata[idx]['place_name']
                sheet.cell(row=idx+2, column=6).value = sensordata[idx]['sensorgroup_name']

                if  sensordata[idx]['sensorgroup_type'] and (sensordata[idx]['sensorgroup_type'] == '0201' or sensordata[idx]['sensorgroup_type'] == '0202' or sensordata[idx]['sensorgroup_type'] == '0204' or sensordata[idx]['sensorgroup_type'] == '0205'):
                    sensor_url = str(rootsplit) +"/ws-02-2-1?sensor_id="+str(sensordata[idx]['sensor_id'])+"&project_id="+str(sensordata[idx]['project_id'])+"&sensorgroup_type="+str(sensordata[idx]['sensorgroup_type'])
                    sheet.cell(row=idx+2, column=7).value = sensor_url
                elif sensordata[idx]['sensorgroup_type'] and sensordata[idx]['sensorgroup_type'] == '0203':
                    if sensordata[idx]['sensor_type'] == 'x':
                        sensor_id = sensordata[idx]['sensor_id']
                        sensorgroup_id = sensordata[idx]['sensorgroup_id']
                        sensor_display_name = sensordata[idx]['sensor_display_name']
                        param = (sensor_display_name, str(sensorgroup_id), str(sensor_id))
                        anorther_sensor = [dict(r) for r in SensorDetailModel.find_by_scatter_another_sensor(param)]
                     
                        another_sensor_id = str(anorther_sensor[0]['sensor_id'])
                        sensor_url = str(rootsplit) +"/ws-02-2-6?sensor_idx="+str(sensordata[idx]['sensor_id'])+"&sensor_idy="+another_sensor_id+"&project_id="+str(sensordata[idx]['project_id'])+"&sensorgroup_type=0203"
                    else:
                        sensor_id = sensordata[idx]['sensor_id']
                        sensorgroup_id = sensordata[idx]['sensorgroup_id']
                        sensor_display_name = sensordata[idx]['sensor_display_name']
                        param = (sensor_display_name, str(sensorgroup_id), str(sensor_id))
                        anorther_sensor = [dict(r) for r in SensorDetailModel.find_by_scatter_another_sensor(param)]
                     
                        another_sensor_id = str(anorther_sensor[0]['sensor_id'])
                        sensor_url = str(rootsplit) +"/ws-02-2-6?sensor_idx="+another_sensor_id+"&sensor_idy="+str(sensordata[idx]['sensor_id'])+"&project_id="+str(sensordata[idx]['project_id'])+"&sensorgroup_type=0203"

                    sheet.cell(row=idx+2, column=7).value = sensor_url
            
  
          


   
        book.remove_sheet(book.get_sheet_by_name('Sheet'))

        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        print(excel_file+file_name)
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }


class fumulalistsensor(Resource):
    parse = reqparse.RequestParser()


    parse.add_argument('function_name', type=str)
    parse.add_argument('sensor_id', type=str)
    parse.add_argument('sensor_fx_check', type=str)
    parse.add_argument('sensor_del_num', type=str)
    parse.add_argument('sensorgroup_type', type=str)
    parse.add_argument('calc', type=str)


    def get(self):

       
        sensor_id = request.args.get('sensor_id')
        data = [dict(r) for r in SensorDetailModel.find_by_sensor_id_function(sensor_id)]
     
        print(data)

        return {'resultCode': '0', "resultString": "SUCCESS", "data":data}, 200

    def post(self):

        params = fumulalistsensor.parse.parse_args()
        print(params)

        sensor_id = params['sensor_id']
        sensor_fx_check = params['sensor_fx_check']

        sensor_datail_obj = SensorDetailModel.find_by_sensor_id(sensor_id)


        sensor_datail_obj.sensor_fx_check = sensor_fx_check
        sensor_datail_obj.save_to_db()
        return {'resultCode': '0', "resultString": "선택하신 공식이 적용 되었습니다."}, 200

    def put(self):

        print("in put formula!!")

        params = fumulalistsensor.parse.parse_args()
        print(params)

        sensor_id = params['sensor_id']
        sensorgroup_type = params['sensorgroup_type']
        sensor_del_num = params['sensor_del_num']
        function_name = params['function_name']
        calc = params['calc']

        function_formula  = params['calc']

        print("!!!!", pi, sin(2*pi), log(10,10), math.e)
        fumula  = params['calc']
        print(fumula)
        # fumula = fumula.replace('$sen', '123')
        fumula = fumula.replace('^', '**')
        fumula = fumula.replace('ln', 'math.log10')
        print(fumula)

        numlist1 = []
        numlist2 = []
        numlist3 = []
        # num = random.randrange(-4000, 4000)
        # numlist.append(num)
        # print(num)
        for i in range(100):
            # print(i)
            num1 = random.uniform(0, 5000)  # 측정값
            num2 = random.uniform(0, 5000)  # 기준값
            num3 = random.uniform(0, 1)     # 가우지팩터
            numlist1.append(num1)
            numlist2.append(num2)
            numlist3.append(num3)
            # ($sen=1234, $ini=1200, Gauge Factor=0.0002
        numlist1.append(2500)
        numlist2.append(2510)
        numlist3.append(0.002)

        try:

            for i, j, w in zip(numlist1, numlist2, numlist3):
             
                fumula01 = fumula
              
                fumula01 = fumula01.replace('$sen', str(i))
                fumula01 = fumula01.replace('$ini', str(j))
                fumula01 = fumula01.replace('$GF', str(w))
               
                result1 = eval(fumula01)


            fumula01 = fumula
              
            fumula01 = fumula01.replace('$sen', str(2500))
            fumula01 = fumula01.replace('$ini', str(2510))
            fumula01 = fumula01.replace('$GF', str(0.002))
            result2 = eval(fumula01)

            print(result2)

            if sensorgroup_type == '0203' :
                print(sensorgroup_type)
                sensor_obj = SensorModel.find_by_id(sensor_id)
                sensor_display_name = sensor_obj.sensor_display_name
                sensorgroup_id = sensor_obj.sensorgroup_id
            
                param = (sensor_display_name, str(sensorgroup_id), sensor_id)

                anorther_sensor = [dict(r) for r in SensorDetailModel.find_by_scatter_another_sensor(param)]

                if len(anorther_sensor) != 0:
                    another_sensor_id = anorther_sensor[0]['sensor_id']
                    another_sensordetail_obj = SensorDetailModel.find_by_sensor_id(another_sensor_id)

                    if sensor_del_num == '1':
                        another_sensordetail_obj.sensor_fx1_name = function_name
                        another_sensordetail_obj.sensor_fx1 = calc

                    elif sensor_del_num == '2':
                        print("2")
                        another_sensordetail_obj.sensor_fx2_name = function_name
                        another_sensordetail_obj.sensor_fx2 = calc
                    
                    elif sensor_del_num == '3':
                        another_sensordetail_obj.sensor_fx3_name = function_name
                        another_sensordetail_obj.sensor_fx3 = calc

                    elif sensor_del_num == '4':
                        another_sensordetail_obj.sensor_fx4_name = function_name
                        another_sensordetail_obj.sensor_fx4 = calc

                    elif sensor_del_num == '5':
                        another_sensordetail_obj.sensor_fx5_name = function_name
                        another_sensordetail_obj.sensor_fx5 = calc

                    another_sensordetail_obj.save_to_db()


            print("sensor_id", sensor_id)
            sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_id)

            if sensor_del_num == '1':
                sensordetail_obj.sensor_fx1_name = function_name
                sensordetail_obj.sensor_fx1 = calc

            elif sensor_del_num == '2':
                print("????")
                sensordetail_obj.sensor_fx2_name = function_name
                sensordetail_obj.sensor_fx2 = calc
            
            elif sensor_del_num == '3':
                sensordetail_obj.sensor_fx3_name = function_name
                sensordetail_obj.sensor_fx3 = calc

            elif sensor_del_num == '4':
                sensordetail_obj.sensor_fx4_name = function_name
                sensordetail_obj.sensor_fx4 = calc

            elif sensor_del_num == '5':
                sensordetail_obj.sensor_fx5_name = function_name
                sensordetail_obj.sensor_fx5 = calc

            sensordetail_obj.save_to_db()


        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "올바른 수식이 아닙니다. 수식을 확인해주세요."}, 500
       
        return {'resultCode': '0', "resultString": "선택하신 공식이 적용 되었습니다."}, 200

    def delete(self):

        params = fumulalistsensor.parse.parse_args()
        print(params)

        sensor_id = params['sensor_id']
        sensorgroup_type = params['sensorgroup_type']
        sensor_del_num = params['sensor_del_num']

        if sensorgroup_type == '0203' :
            print(sensorgroup_type)
            sensor_obj = SensorModel.find_by_id(sensor_id)
            sensor_display_name = sensor_obj.sensor_display_name
            sensorgroup_id = sensor_obj.sensorgroup_id
           
            param = (sensor_display_name, str(sensorgroup_id), sensor_id)

            anorther_sensor = [dict(r) for r in SensorDetailModel.find_by_scatter_another_sensor(param)]

            if len(anorther_sensor) != 0:
                another_sensor_id = anorther_sensor[0]['sensor_id']
                another_sensordetail_obj = SensorDetailModel.find_by_sensor_id(another_sensor_id)

                if sensor_del_num == '1':
                    another_sensordetail_obj.sensor_fx1_name = None
                    another_sensordetail_obj.sensor_fx1 = None

                elif sensor_del_num == '2':
                    print("2")
                    another_sensordetail_obj.sensor_fx2_name = None
                    another_sensordetail_obj.sensor_fx2 = None
                
                elif sensor_del_num == '3':
                    another_sensordetail_obj.sensor_fx3_name = None
                    another_sensordetail_obj.sensor_fx3 = None

                elif sensor_del_num == '4':
                    another_sensordetail_obj.sensor_fx4_name = None
                    another_sensordetail_obj.sensor_fx4 = None

                elif sensor_del_num == '5':
                    another_sensordetail_obj.sensor_fx5_name = None
                    another_sensordetail_obj.sensor_fx5 = None

                another_sensordetail_obj.save_to_db()


        print("sensor_id", sensor_id)
        sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_id)

        if sensor_del_num == '1':
            sensordetail_obj.sensor_fx1_name = None
            sensordetail_obj.sensor_fx1 = None

        elif sensor_del_num == '2':
            print("????")
            sensordetail_obj.sensor_fx2_name = None
            sensordetail_obj.sensor_fx2 = None
        
        elif sensor_del_num == '3':
            sensordetail_obj.sensor_fx3_name = None
            sensordetail_obj.sensor_fx3 = None

        elif sensor_del_num == '4':
            sensordetail_obj.sensor_fx4_name = None
            sensordetail_obj.sensor_fx4 = None

        elif sensor_del_num == '5':
            sensordetail_obj.sensor_fx5_name = None
            sensordetail_obj.sensor_fx5 = None

        sensordetail_obj.save_to_db()
       
        return {'resultCode': '0', "resultString": "선택하신 공식이 적용 되었습니다."}, 200


        

class Sensorgaugemodalscatter(Resource):
    parse = reqparse.RequestParser()
 

    parse.add_argument('sensor_idx', type=str)
    parse.add_argument('sensor_idy', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_namex', type=str)
    parse.add_argument('sensor_namey', type=str)

    parse.add_argument('gauge_x', type=str)
    parse.add_argument('gauge_y', type=str)


    def post(self):

        print("들어옴")
        params = Sensorgaugemodalscatter.parse.parse_args()
        print(params)

        sensor_idx = params['sensor_idx']
        sensor_idy = params['sensor_idy']

        datarogger_id = params['datarogger_id']
        gauge_x = params['gauge_x']
        gauge_y = params['gauge_y']

        sensor_namex = params['sensor_namex']
        sensor_namey = params['sensor_namey']

       
   
      

        # print(jsonData)
        sensor_detail_idx = SensorDetailModel.find_by_sensor_id(sensor_idx).sensor_detail_id
        sensor_detail_idy = SensorDetailModel.find_by_sensor_id(sensor_idy).sensor_detail_id
    
        sensor_detail_objx = SensorDetailModel.find_by_id(sensor_detail_idx)
        sensor_detail_objy = SensorDetailModel.find_by_id(sensor_detail_idy)

        sensor_detail_objx.sensor_gauge_factor = gauge_x
        sensor_detail_objy.sensor_gauge_factor = gauge_y

        try:
            sensor_detail_objx.save_to_db()
            sensor_detail_objy.save_to_db()
          

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

    
        return {'resultCode': '0', "resultString": "SUCCESS"}, 200


class Sensordetailall(Resource):
    parse = reqparse.RequestParser()

    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('datarogger_id', type=str)
    parse.add_argument('sensor_gl1_max', type=str)
    parse.add_argument('sensor_gl1_min', type=str)
    parse.add_argument('sensor_gl2_max', type=str)
    parse.add_argument('sensor_gl2_min', type=str)
    parse.add_argument('sensor_gl3_max', type=str)
    parse.add_argument('sensor_gl3_min', type=str)
    parse.add_argument('sensor_fx1', type=str)
    parse.add_argument('sensor_fx2', type=str)
    parse.add_argument('sensor_fx3', type=str)
    parse.add_argument('sensor_fx4', type=str)
    parse.add_argument('sensor_fx5', type=str)
    parse.add_argument('sensor_max', type=str)
    parse.add_argument('sensor_min', type=str)
    
    parse.add_argument('sensor_fx1_name', type=str)
    parse.add_argument('sensor_fx2_name', type=str)
    parse.add_argument('sensor_fx3_name', type=str)
    parse.add_argument('sensor_fx4_name', type=str)
    parse.add_argument('sensor_fx5_name', type=str)

    parse.add_argument('sensor_fx1_id', type=str)
    parse.add_argument('sensor_fx2_id', type=str)
    parse.add_argument('sensor_fx3_id', type=str)
    parse.add_argument('sensor_fx4_id', type=str)
    parse.add_argument('sensor_fx5_id', type=str)

    parse.add_argument('sensor_weight', type=str)
    parse.add_argument('sensor_deviation', type=str)
    parse.add_argument('sensor_st_over_ex', type=str)
    parse.add_argument('sensor_st_over_wt', type=str)
    parse.add_argument('sensor_dev_over_ex', type=str)
    parse.add_argument('sensor_dev_over_wt', type=str)
    parse.add_argument('sensor_null_ex', type=str)
    parse.add_argument('sensor_default_wt', type=str)
    parse.add_argument('sensor_noti', type=str)

    parse.add_argument('sensor_detail_id', type=str)
    parse.add_argument('sensor_id', type=str)
    parse.add_argument('sensorgroup_id', type=str)
    parse.add_argument('sensor_id_list', type=str)



    def post(self):

        params = Sensordetailall.parse.parse_args()
        print(params)

        # sensor_id = request.args.get('sensor_id')
        # datarogger_id = request.args.get('datarogger_id')

        print("!!!!!!!!!!!!!!")

        sensor_detail_id = params['sensor_detail_id']
        sensor_gl1_max = params['sensor_gl1_max']
        sensor_gl1_min = params['sensor_gl1_min']
        sensor_gl2_max = params['sensor_gl2_max']
        sensor_gl2_min = params['sensor_gl2_min']
        sensor_gl3_max = params['sensor_gl3_max']
        sensor_gl3_min = params['sensor_gl3_min']
        sensor_fx1 = params['sensor_fx1']
        sensor_fx2 = params['sensor_fx2']
        sensor_fx3 = params['sensor_fx3']
        sensor_fx4 = params['sensor_fx4']
        sensor_fx5 = params['sensor_fx5']
        sensor_max = params['sensor_max']
        sensor_min = params['sensor_min']

        sensor_fx1_name = params['sensor_fx1_name']
        sensor_fx2_name = params['sensor_fx2_name']
        sensor_fx3_name = params['sensor_fx3_name']
        sensor_fx4_name = params['sensor_fx4_name']
        sensor_fx5_name = params['sensor_fx5_name']
        
        sensor_fx1_id = params['sensor_fx1_id']
        sensor_fx2_id = params['sensor_fx2_id']
        sensor_fx3_id = params['sensor_fx3_id']
        sensor_fx4_id = params['sensor_fx4_id']
        sensor_fx5_id = params['sensor_fx5_id']
        
        sensor_weight = params['sensor_weight']
        sensor_deviation = params['sensor_deviation']

        sensor_st_over_ex = params['sensor_st_over_ex']
        sensor_st_over_wt = params['sensor_st_over_wt']
        sensor_dev_over_ex = params['sensor_dev_over_ex']
        sensor_dev_over_wt = params['sensor_dev_over_wt']
        sensor_null_ex = params['sensor_null_ex']
        sensor_default_wt = params['sensor_default_wt']
        
        sensor_noti = params['sensor_noti']
        sensor_id = params['sensor_id']
        sensorgroup_id = params['sensorgroup_id']
        datarogger_id = params['datarogger_id']

        sensor_id_list = params['sensor_id_list']
        sensor_id_list = sensor_id_list.split(',')[0:-1]

        print(sensor_id_list)
        

        if len(sensor_gl1_max) ==0: sensor_gl1_max = None
        if len(sensor_gl1_min) ==0: sensor_gl1_min = None
        if len(sensor_gl2_max) ==0: sensor_gl2_max = None
        if len(sensor_gl2_min) ==0: sensor_gl2_min = None
        if len(sensor_gl3_max) ==0: sensor_gl3_max = None
        if len(sensor_gl3_min) ==0: sensor_gl3_min = None
        # if len(sensor_fx1) ==0: sensor_fx1 = None
        # if len(sensor_fx2) ==0: sensor_fx2 = None
        # if len(sensor_fx3) ==0: sensor_fx3 = None
        # if len(sensor_fx4) ==0: sensor_fx4 = None
        # if len(sensor_fx5) ==0: sensor_fx5 = None
        if len(sensor_max) ==0: sensor_max = None
        if len(sensor_min) ==0: sensor_min = None
        if len(sensor_weight) ==0: sensor_weight = None
        if len(sensor_deviation) ==0: sensor_deviation = None
        if len(sensor_st_over_ex) ==0: sensor_st_over_ex = None
        if len(sensor_st_over_wt) ==0: sensor_st_over_wt = None
        if len(sensor_dev_over_ex) ==0: sensor_dev_over_ex = None
        if len(sensor_dev_over_wt) ==0: sensor_dev_over_wt = None
        if len(sensor_null_ex) ==0: sensor_null_ex = None
        if len(sensor_default_wt) ==0: sensor_default_wt = None
        if len(sensor_noti) ==0: sensor_noti = None
  

        use_yn = 'Y'
        user_id = current_user.user_id
        create_date = datetime.now()
        modify_date = datetime.now()
     
        print("sensor_noti", sensor_noti)

        if sensorgroup_id == "선택":
            sensorgroup_id = None
        if sensor_fx1_name == "선택":
            sensor_fx1_name = None
        if sensor_fx2_name == "선택":
            sensor_fx2_name = None
        if sensor_fx3_name == "선택":
            sensor_fx3_name = None
        if sensor_fx4_name == "선택":
            sensor_fx4_name = None
        if sensor_fx5_name == "선택":
            sensor_fx5_name = None

        if sensor_fx1 == "undefined":
            sensor_fx1 = None
        if sensor_fx2 == "undefined":
            sensor_fx2 = None
        if sensor_fx3 == "undefined":
            sensor_fx3 = None
        if sensor_fx4 == "undefined":
            sensor_fx4 = None
        if sensor_fx5 == "undefined":
            sensor_fx5 = None

        if len(sensor_fx1) !=0: sensor_fx_check = '1'
        elif len(sensor_fx2) != 0: sensor_fx_check = '2'
        elif len(sensor_fx3) != 0: sensor_fx_check = '3'
        elif len(sensor_fx4) != 0: sensor_fx_check = '4'
        elif len(sensor_fx5) != 0: sensor_fx_check = '5'
        else : sensor_fx_check = None


       
        if sensor_noti == 'null': 
            print("!!@#!@#")
            sensor_noti = None
        

        


        

        
        try:

            for i in range(len(sensor_id_list)):
                sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_id_list[i])

                if sensordetail_obj:
                    sensor_detail_id = sensordetail_obj.sensor_detail_id
                    print("sensordetail id")
                    

                    sensordetail_obj.sensor_gl1_max = sensor_gl1_max
                    sensordetail_obj.sensor_gl1_min = sensor_gl1_min
                    sensordetail_obj.sensor_gl2_max = sensor_gl2_max
                    sensordetail_obj.sensor_gl2_min = sensor_gl2_min
                    sensordetail_obj.sensor_gl3_max = sensor_gl3_max
                    sensordetail_obj.sensor_gl3_min = sensor_gl3_min


                    if sensordetail_obj.sensor_fx1_id != sensor_fx1_id :
                        sensordetail_obj.sensor_fx1 = sensor_fx1
                        sensordetail_obj.sensor_fx1_id = sensor_fx1_id
                        sensordetail_obj.sensor_fx1_name = sensor_fx1_name

                    if sensordetail_obj.sensor_fx2_id != sensor_fx2_id :
                        sensordetail_obj.sensor_fx2 = sensor_fx2
                        sensordetail_obj.sensor_fx2_id = sensor_fx2_id
                        sensordetail_obj.sensor_fx2_name = sensor_fx2_name

                    if sensordetail_obj.sensor_fx3_id != sensor_fx3_id :
                        sensordetail_obj.sensor_fx3 = sensor_fx3
                        sensordetail_obj.sensor_fx3_id = sensor_fx3_id
                        sensordetail_obj.sensor_fx3_name = sensor_fx3_name

                    if sensordetail_obj.sensor_fx4_id != sensor_fx4_id :
                        sensordetail_obj.sensor_fx4 = sensor_fx4
                        sensordetail_obj.sensor_fx4_id = sensor_fx4_id
                        sensordetail_obj.sensor_fx4_name = sensor_fx4_name

                    if sensordetail_obj.sensor_fx5_id != sensor_fx5_id :
                        sensordetail_obj.sensor_fx5 = sensor_fx5
                        sensordetail_obj.sensor_fx5_id = sensor_fx5_id
                        sensordetail_obj.sensor_fx5_name = sensor_fx5_name

        
                    sensordetail_obj.sensor_fx_check = sensor_fx_check

                    sensordetail_obj.sensor_max = sensor_max
                    sensordetail_obj.sensor_min = sensor_min
                    sensordetail_obj.sensor_weight = sensor_weight
                    sensordetail_obj.sensor_deviation = sensor_deviation

                    sensordetail_obj.sensor_st_over_ex = sensor_st_over_ex
                    sensordetail_obj.sensor_st_over_wt = sensor_st_over_wt
                    sensordetail_obj.sensor_dev_over_ex = sensor_dev_over_ex
                    sensordetail_obj.sensor_dev_over_wt = sensor_dev_over_wt
                    sensordetail_obj.sensor_null_ex = sensor_null_ex
                    sensordetail_obj.sensor_default_wt = sensor_default_wt
                    sensordetail_obj.sensor_noti = sensor_noti

                    sensordetail_obj.save_to_db()

                    if sensorgroup_id:

                        sensor_obj = SensorModel.find_by_id(sensor_id_list[i])
                        print(sensor_obj.sensorgroup_id)
                        sensor_obj.sensorgroup_id = sensorgroup_id
                        sensor_obj.save_to_db()

                        sensordetail_obj = SensorDetailModel.find_by_sensor_id(sensor_id_list[i])
                        sensor_initial_date = sensordetail_obj.sensor_initial_date

                        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
                        sensorgroup_obj.sensorgroup_initial_date = sensor_initial_date
                        sensorgroup_obj.save_to_db()

                else:
          
                    print("not sensordetail id")
                    sensor_obj = SensorModel.find_by_id(sensor_id_list[i])
                    sensor_name = sensor_obj.sensor_name
                    datarogger_id = sensor_obj.datarogger_id
                    data = [dict(r) for r in RawdataModel.sensor_data_first(str(datarogger_id), str(sensor_name))]
                    print(data)

                    if(len(data) != 0):
                        sensor_initial_data = data[0]['sensor_data']
                        sensor_initial_date = data[0]['sensor_data_date']

                    else: 
                        print("not data")
                        sensor_initial_data = None
                        sensor_initial_date = None
                    print("없어요!!!!")

                    sensordetail_obj = SensorDetailModel(sensor_initial_date,sensor_initial_data, sensor_gl1_max, sensor_gl1_min, sensor_gl2_max, sensor_gl2_min, sensor_gl3_max, sensor_gl3_min, sensor_fx1, sensor_fx2, sensor_fx3, sensor_fx4, sensor_fx5,
                                sensor_max, sensor_min, sensor_weight, sensor_deviation, sensor_st_over_ex, sensor_st_over_wt, sensor_dev_over_ex, sensor_dev_over_wt, sensor_null_ex, sensor_default_wt, sensor_noti, '0.002', sensor_fx1_name,  sensor_fx2_name,  sensor_fx3_name,  sensor_fx4_name,  sensor_fx5_name,  sensor_fx_check,  
                                sensor_fx1_id, sensor_fx2_id, sensor_fx3_id, sensor_fx4_id, sensor_fx5_id, use_yn, user_id,
                                sensor_id_list[i], create_date, modify_date)



                    if sensorgroup_id:

                        sensor_obj = SensorModel.find_by_id(sensor_id_list[i])
                        print(sensor_obj.sensorgroup_id)
                        sensor_obj.sensorgroup_id = sensorgroup_id
                        sensor_obj.save_to_db()

                        sensorgroup_obj = SensorgroupModel.find_by_id(sensorgroup_id)
                        sensorgroup_obj.sensorgroup_initial_date = sensor_initial_date
                        sensorgroup_obj.save_to_db()
            
                    sensordetail_obj.save_to_db()
                

    
                
          
                
                

        except Exception as e:

            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 500

    
        return {'resultCode': '0', "resultString": "SUCCESS"}, 200